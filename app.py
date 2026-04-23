from __future__ import annotations

import os
import re
import secrets
import smtplib
import sqlite3
import subprocess
import ssl
from datetime import datetime, timedelta
from functools import wraps
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from werkzeug.middleware.proxy_fix import ProxyFix

import csv
import shutil
from email.message import EmailMessage
from io import BytesIO, StringIO

from flask import Flask, abort, flash, g, redirect, render_template, request, send_file, session, url_for
from flask_login import LoginManager, UserMixin, current_user, login_required, login_user, logout_user
from werkzeug.security import check_password_hash, generate_password_hash

try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except Exception:
    psycopg2 = None
    RealDictCursor = None

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

BASE_DIR = Path(__file__).resolve().parent
PROD_DB_PATH = BASE_DIR / 'instance' / 'contabilidad_condominio.db'
DEMO_DB_PATH = BASE_DIR / 'instance' / 'contabilidad_condominio_demo.db'
DB_PATH = PROD_DB_PATH
INSTANCE_DIR = BASE_DIR / 'instance'
BACKUP_DIR = BASE_DIR / 'backups'
INSTANCE_DIR.mkdir(exist_ok=True)
BACKUP_DIR.mkdir(exist_ok=True)

SCHOOL_NAME = 'Parcelia'
SCHOOL_LOCATION = 'Administración de parcelas en condominios'
DEFAULT_CONDOMINIO_NAME = 'Condominio Base'
ALLOWED_ROLES = ('admin', 'presidente', 'tesorero', 'secretario', 'comite', 'propietario')
ACTA_ESTADOS = ('borrador', 'en_revision', 'aprobada')
ROLE_LABELS = {
    'admin': 'Administrador',
    'presidente': 'Presidente',
    'tesorero': 'Tesorero',
    'secretario': 'Secretario',
    'comite': 'Comité',
    'propietario': 'Propietario',
}


class User(UserMixin):
    def __init__(self, row: Any, is_demo_db: bool = False):
        self.id = str(row['id'])
        self.username = row['username']
        self.password_hash = row['password_hash']
        self.role = row['role']
        self.nombre = row['nombre']
        self.activo = bool(row['activo'])
        self.condominio_id = row['condominio_id'] if 'condominio_id' in row.keys() else None
        self.condominio_nombre = row['condominio_nombre'] if 'condominio_nombre' in row.keys() else None
        self.parcela_id = row['parcela_id'] if 'parcela_id' in row.keys() else None
        self.is_demo_db = is_demo_db
        self.must_change_password = bool(row['must_change_password']) if 'must_change_password' in row.keys() else False


    def get_id(self) -> str:
        prefix = 'demo' if self.is_demo_db else 'prod'
        return f'{prefix}:{self.id}'

    def can_edit(self) -> bool:
        return self.role in ('admin', 'presidente', 'tesorero', 'secretario')

    def can_delete(self) -> bool:
        return self.role in ('admin', 'presidente', 'tesorero')

    def can_manage_finance(self) -> bool:
        return self.role in ('admin', 'presidente', 'tesorero')

    def can_manage_actas(self) -> bool:
        return self.role in ('admin', 'presidente', 'secretario')

    def is_admin(self) -> bool:
        return self.role == 'admin'

    def is_global_admin(self) -> bool:
        return self.role == 'admin' and self.condominio_id is None

    def can_manage_condominio(self, condominio_id: int | None) -> bool:
        if self.is_global_admin():
            return True
        return self.role == 'admin' and condominio_id is not None and self.condominio_id == condominio_id

    @property
    def role_label(self) -> str:
        return ROLE_LABELS.get(self.role, self.role.title())

    def can_view_only(self) -> bool:
        return self.role == 'propietario'

    def needs_password_change(self) -> bool:
        return self.must_change_password


class DBAdapter:
    def __init__(self, url: str):
        self.url = url
        self.kind = 'postgres' if url.startswith('postgresql://') or url.startswith('postgres://') else 'sqlite'
        self.conn = self._connect()

    def _connect(self):
        if self.kind == 'sqlite':
            conn = sqlite3.connect(self.url)
            conn.row_factory = sqlite3.Row
            conn.execute('PRAGMA foreign_keys = ON')
            return conn
        if psycopg2 is None:
            raise RuntimeError('psycopg2-binary no está instalado. Ejecuta: python3 -m pip install psycopg2-binary')
        conn = psycopg2.connect(self.url, cursor_factory=RealDictCursor)
        conn.autocommit = False
        return conn

    def close(self):
        self.conn.close()

    def commit(self):
        self.conn.commit()

    def rollback(self):
        self.conn.rollback()

    def cursor(self):
        return self.conn.cursor()

    def _convert_sql(self, sql: str) -> str:
        if self.kind == 'sqlite':
            return sql
        return re.sub(r'\?', '%s', sql)

    def execute(self, sql: str, params: tuple | list | None = None):
        cur = self.conn.cursor()
        cur.execute(self._convert_sql(sql), params or [])
        return cur

    def executescript(self, script: str):
        if self.kind == 'sqlite':
            return self.conn.executescript(script)
        statements = [s.strip() for s in script.split(';') if s.strip()]
        cur = self.conn.cursor()
        for stmt in statements:
            cur.execute(stmt)
        return cur

    def fetchone(self, sql: str, params: tuple | list | None = None):
        cur = self.execute(sql, params)
        return cur.fetchone()

    def fetchall(self, sql: str, params: tuple | list | None = None):
        cur = self.execute(sql, params)
        return cur.fetchall()



def table_exists(db: DBAdapter, table_name: str) -> bool:
    try:
        if db.kind == 'sqlite':
            row = db.fetchone("SELECT name FROM sqlite_master WHERE type IN ('table', 'view') AND name = ?", (table_name,))
        else:
            row = db.fetchone("SELECT table_name FROM information_schema.tables WHERE table_schema = 'public' AND table_name = ?", (table_name,))
        return row is not None
    except Exception:
        return False


def column_exists(db: DBAdapter, table_name: str, column_name: str) -> bool:
    try:
        if db.kind == 'sqlite':
            rows = db.fetchall(f'PRAGMA table_info({table_name})')
            return any(r['name'] == column_name for r in rows)
        row = db.fetchone(
            "SELECT column_name FROM information_schema.columns WHERE table_schema = 'public' AND table_name = ? AND column_name = ?",
            (table_name, column_name),
        )
        return row is not None
    except Exception:
        return False


def migrate_legacy_parcelas_schema(db: DBAdapter) -> None:
    """Migra instalaciones antiguas basadas en alumnos/pagos_alumnos al esquema parcelas/pagos_parcelas."""
    try:
        has_old_parcelas = table_exists(db, 'alumnos')
        has_new_parcelas = table_exists(db, 'parcelas')
        if has_old_parcelas and has_new_parcelas:
            old_count = db.fetchone('SELECT COUNT(*) AS total FROM alumnos')
            new_count = db.fetchone('SELECT COUNT(*) AS total FROM parcelas')
            if int((old_count['total'] if old_count else 0) or 0) > 0 and int((new_count['total'] if new_count else 0) or 0) == 0:
                common = ['id', 'nombre', 'curso', 'cuota_mensual', 'apoderado', 'telefono', 'direccion', 'observacion_ficha', 'activo']
                if column_exists(db, 'alumnos', 'condominio_id') and column_exists(db, 'parcelas', 'condominio_id'):
                    common.append('condominio_id')
                cols = ', '.join(common)
                db.execute(f'INSERT INTO parcelas ({cols}) SELECT {cols} FROM alumnos')
                db.commit()
    except Exception:
        db.rollback()

    try:
        has_old_pagos = table_exists(db, 'pagos_alumnos')
        has_new_pagos = table_exists(db, 'pagos_parcelas')
        if has_old_pagos and has_new_pagos:
            old_count = db.fetchone('SELECT COUNT(*) AS total FROM pagos_alumnos')
            new_count = db.fetchone('SELECT COUNT(*) AS total FROM pagos_parcelas')
            if int((old_count['total'] if old_count else 0) or 0) > 0 and int((new_count['total'] if new_count else 0) or 0) == 0:
                source_fk = 'parcela_id' if column_exists(db, 'pagos_alumnos', 'parcela_id') else 'alumno_id'
                common = [('id', 'id'), (source_fk, 'parcela_id'), ('fecha', 'fecha'), ('mes', 'mes'), ('monto', 'monto')]
                for legacy_col, new_col in [('observacion', 'observacion'), ('movimiento_id', 'movimiento_id'), ('condominio_id', 'condominio_id')]:
                    if column_exists(db, 'pagos_alumnos', legacy_col) and column_exists(db, 'pagos_parcelas', new_col):
                        common.append((legacy_col, new_col))
                src = ', '.join(c[0] for c in common)
                dst = ', '.join(c[1] for c in common)
                db.execute(f'INSERT INTO pagos_parcelas ({dst}) SELECT {src} FROM pagos_alumnos')
                db.commit()
    except Exception:
        db.rollback()

    try:
        if column_exists(db, 'movimientos', 'alumno_id') and column_exists(db, 'movimientos', 'parcela_id'):
            db.execute('UPDATE movimientos SET parcela_id = alumno_id WHERE parcela_id IS NULL AND alumno_id IS NOT NULL')
            db.commit()
    except Exception:
        db.rollback()

    # Compatibilidad de solo lectura para herramientas antiguas o scripts externos.
    if db.kind == 'sqlite':
        statements = [
            'DROP VIEW IF EXISTS alumnos',
            'DROP VIEW IF EXISTS pagos_alumnos',
            'CREATE VIEW IF NOT EXISTS alumnos AS SELECT * FROM parcelas',
            'CREATE VIEW IF NOT EXISTS pagos_alumnos AS SELECT id, parcela_id AS alumno_id, fecha, mes, monto, observacion, movimiento_id, condominio_id FROM pagos_parcelas',
        ]
        for stmt in statements:
            try:
                db.execute(stmt)
                db.commit()
            except Exception:
                db.rollback()


def create_app() -> Flask:
    app = Flask(__name__)
    debug_enabled = os.environ.get('APP_DEBUG', '0') == '1'
    secret_key = os.environ.get('SECRET_KEY', '').strip()
    if not secret_key:
        if debug_enabled:
            secret_key = 'dev-inseguro-cambia-esta-clave'
        else:
            raise RuntimeError('SECRET_KEY no está configurada. Define una clave segura antes de producción.')
    if not debug_enabled and secret_key in {'cambia-esta-clave', 'dev-inseguro-cambia-esta-clave'}:
        raise RuntimeError('SECRET_KEY insegura. Define una clave distinta para producción.')

    app.config['SECRET_KEY'] = secret_key
    app.config['DATABASE'] = os.environ.get('DATABASE_URL', str(DB_PATH))
    app.config['MAIL_TO'] = os.environ.get('MAIL_TO', 'parceliaMP@gmail.com').strip() or 'parceliaMP@gmail.com'
    app.config['SMTP_HOST'] = os.environ.get('SMTP_HOST', 'smtp.gmail.com').strip() or 'smtp.gmail.com'
    app.config['SMTP_PORT'] = int(os.environ.get('SMTP_PORT', '587'))
    app.config['SMTP_USER'] = os.environ.get('SMTP_USER', '').strip()
    app.config['SMTP_PASSWORD'] = os.environ.get('SMTP_PASSWORD', '').strip()
    app.config['SMTP_USE_TLS'] = os.environ.get('SMTP_USE_TLS', '1') == '1'
    app.config['SMTP_USE_SSL'] = os.environ.get('SMTP_USE_SSL', '0') == '1'
    app.config['MAIL_FROM'] = os.environ.get('MAIL_FROM', app.config['SMTP_USER'] or app.config['MAIL_TO']).strip()

    app_env = os.environ.get('APP_ENV', os.environ.get('FLASK_ENV', '')).strip().lower()
    explicit_secure = os.environ.get('SESSION_COOKIE_SECURE')
    if explicit_secure is None:
        secure_cookies = app_env in {'prod', 'production'} or os.environ.get('RENDER', '').lower() == 'true'
    else:
        secure_cookies = explicit_secure == '1'

    explicit_remember_secure = os.environ.get('REMEMBER_COOKIE_SECURE')
    if explicit_remember_secure is None:
        remember_secure = secure_cookies
    else:
        remember_secure = explicit_remember_secure == '1'

    app.config['PREFERRED_URL_SCHEME'] = 'https' if secure_cookies else 'http'
    app.config['SESSION_COOKIE_SECURE'] = secure_cookies
    app.config['SESSION_COOKIE_HTTPONLY'] = True
    app.config['SESSION_COOKIE_SAMESITE'] = os.environ.get('SESSION_COOKIE_SAMESITE', 'Lax')
    app.config['REMEMBER_COOKIE_SECURE'] = remember_secure
    app.config['REMEMBER_COOKIE_HTTPONLY'] = True
    app.config['REMEMBER_COOKIE_SAMESITE'] = os.environ.get('REMEMBER_COOKIE_SAMESITE', 'Lax')
    session_lifetime_minutes = int(os.environ.get('SESSION_LIFETIME_MINUTES', '720'))
    inactivity_timeout_minutes = int(os.environ.get('INACTIVITY_TIMEOUT_MINUTES', '15'))
    app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=session_lifetime_minutes)
    app.config['INACTIVITY_TIMEOUT_SECONDS'] = inactivity_timeout_minutes * 60
    app.config['MAX_CONTENT_LENGTH'] = int(os.environ.get('MAX_CONTENT_LENGTH_MB', '16')) * 1024 * 1024
    app.config['API_TOKEN_MAX_AGE_SECONDS'] = int(os.environ.get('API_TOKEN_MAX_AGE_SECONDS', str(60 * 60 * 24 * 7)))
    app.config['API_ALLOWED_ORIGINS'] = [o.strip() for o in os.environ.get('API_ALLOWED_ORIGINS', '*').split(',') if o.strip()]

    if os.environ.get('TRUST_PROXY', '1') == '1':
        app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)

    login_manager = LoginManager()
    login_manager.login_view = 'login'
    login_manager.login_message = 'Inicia sesión para continuar.'
    login_manager.login_message_category = 'warning'
    login_manager.init_app(app)

    def api_serializer() -> URLSafeTimedSerializer:
        return URLSafeTimedSerializer(app.config['SECRET_KEY'], salt='parcelia-mobile-api')

    def generate_api_token(user: User) -> str:
        return api_serializer().dumps({
            'uid': int(user.id.split(':')[-1]) if ':' in user.id else int(user.id),
            'mode': 'demo' if getattr(user, 'is_demo_db', False) else 'prod',
            'role': user.role,
        })

    def decode_api_token(token: str) -> dict[str, Any]:
        max_age = int(app.config.get('API_TOKEN_MAX_AGE_SECONDS', 0) or 0)
        kwargs = {'max_age': max_age} if max_age > 0 else {}
        return api_serializer().loads(token, **kwargs)

    def generate_csrf_token() -> str:
        token = session.get('_csrf_token')
        if not token:
            token = secrets.token_urlsafe(32)
            session['_csrf_token'] = token
        return token

    def validate_csrf() -> None:
        if request.path.startswith('/api/'):
            return
        if request.method in {'GET', 'HEAD', 'OPTIONS', 'TRACE'}:
            return
        provided_token = (
            request.form.get('_csrf_token', '')
            or request.headers.get('X-CSRFToken', '')
            or request.headers.get('X-CSRF-Token', '')
        ).strip()
        expected_token = session.get('_csrf_token', '')
        if not expected_token or not provided_token or not secrets.compare_digest(provided_token, expected_token):
            abort(400)

    @app.before_request
    def enforce_request_security() -> None:
        if request.method not in {'GET', 'HEAD', 'OPTIONS', 'TRACE'}:
            validate_csrf()
        if current_user.is_authenticated:
            session.permanent = True
            now_ts = int(datetime.now().timestamp())
            last_activity_ts = int(session.get('last_activity_ts', now_ts) or now_ts)
            inactivity_timeout = int(app.config.get('INACTIVITY_TIMEOUT_SECONDS', 0) or 0)
            ignored_endpoints = {'static', 'login', 'logout', 'healthz'}
            if (
                inactivity_timeout > 0
                and request.endpoint not in ignored_endpoints
                and now_ts - last_activity_ts > inactivity_timeout
            ):
                logout_user()
                session.clear()
                flash('Tu sesión expiró por inactividad.', 'warning')
                return redirect(url_for('login', next=request.full_path if request.full_path.startswith('/') else request.path))
            if getattr(current_user, 'needs_password_change', lambda: False)():
                allowed_endpoints = {'first_login_password', 'logout', 'static'}
                if request.endpoint not in allowed_endpoints:
                    flash('Primero debes crear tu nueva contraseña para continuar.', 'warning')
                    return redirect(url_for('first_login_password'))
            session['last_activity_ts'] = now_ts

    @app.after_request
    def set_security_headers(response):
        response.headers.setdefault('X-Frame-Options', 'SAMEORIGIN')
        response.headers.setdefault('X-Content-Type-Options', 'nosniff')
        response.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
        response.headers.setdefault('Permissions-Policy', 'camera=(), microphone=(), geolocation=()')
        if request.is_secure or request.headers.get('X-Forwarded-Proto', 'http') == 'https':
            response.headers.setdefault('Strict-Transport-Security', 'max-age=31536000; includeSubDomains')
        return response

    def is_safe_redirect_target(target: str | None) -> bool:
        if not target:
            return False
        parsed_target = urlparse(target)
        if parsed_target.scheme or parsed_target.netloc:
            ref_url = urlparse(request.host_url)
            return (
                parsed_target.scheme in {'http', 'https'}
                and parsed_target.netloc == ref_url.netloc
            )
        return target.startswith('/') and not target.startswith('//')

    def redirect_to_local_url(target: str | None, fallback_endpoint: str, **fallback_values):
        if is_safe_redirect_target(target):
            return redirect(target)
        return redirect(url_for(fallback_endpoint, **fallback_values))

    def redirect_back(fallback_endpoint: str, **fallback_values):
        return redirect_to_local_url(request.referrer, fallback_endpoint, **fallback_values)

    @login_manager.unauthorized_handler
    def unauthorized():
        if request.path.startswith('/api/'):
            return {'ok': False, 'error': 'auth_required', 'message': 'Debes autenticarte para usar la API.'}, 401
        flash('Inicia sesión para continuar.', 'warning')
        return redirect(url_for('login', next=request.full_path if request.full_path.startswith('/') else request.path))

    @app.context_processor
    def inject_globals() -> dict[str, Any]:
        db = get_db()
        condominio_actual = get_current_condominio(db)
        return {
            'csrf_token': generate_csrf_token,
            'SCHOOL_NAME': SCHOOL_NAME,
            'SCHOOL_LOCATION': (dict(condominio_actual) if condominio_actual else {}).get('nombre', 'Sin condominio seleccionado'),
            'now': datetime.now(),
            'formato_monto': formato_monto,
            'estado_cuota': estado_cuota,
            'current_user': current_user,
            'backup_dir': BACKUP_DIR,
            'db_engine': 'PostgreSQL' if is_postgres_url(app.config['DATABASE']) else 'SQLite',
            'ROLE_LABELS': ROLE_LABELS,
            'current_condominio': condominio_actual,
            'active_condominio_id': get_current_condominio_id(db),
            'admin_condominios': db.fetchall('SELECT id, nombre, activo FROM condominios ORDER BY nombre') if getattr(current_user, 'is_authenticated', False) and getattr(current_user, 'is_global_admin', lambda: False)() else [],
            'inactivity_timeout_seconds': int(app.config.get('INACTIVITY_TIMEOUT_SECONDS', 0) or 0),
        }

    def get_database_url_for_request() -> str:
        return str(DEMO_DB_PATH) if session.get('db_mode') == 'demo' else app.config['DATABASE']

    def get_db() -> DBAdapter:
        selected_url = get_database_url_for_request()
        if 'db' not in g or getattr(g, 'db_url', None) != selected_url:
            old_db = g.pop('db', None)
            if old_db is not None:
                old_db.close()
            g.db = DBAdapter(selected_url)
            g.db_url = selected_url
        return g.db


    @app.url_defaults
    def add_condominio_query(endpoint: str, values: dict[str, Any]) -> None:
        try:
            if not getattr(current_user, 'is_authenticated', False):
                return
            if not getattr(current_user, 'is_global_admin', lambda: False)():
                return
            if 'condominio_id' in values:
                return
            selected = request.args.get('condominio_id', '').strip()
            if selected.isdigit():
                values['condominio_id'] = int(selected)
        except Exception:
            return

    @login_manager.user_loader
    def load_user(user_id: str):
        mode = 'prod'
        numeric_user_id = user_id
        if ':' in user_id:
            mode, numeric_user_id = user_id.split(':', 1)
        if mode not in ('prod', 'demo'):
            mode = 'prod'
        session['db_mode'] = mode
        db = get_db()
        row = db.fetchone('SELECT u.*, c.nombre AS condominio_nombre FROM usuarios u LEFT JOIN condominios c ON c.id = u.condominio_id WHERE u.id = ? AND u.activo = 1', (int(numeric_user_id),))
        return User(row, is_demo_db=(mode == 'demo')) if row else None

    @app.teardown_appcontext
    def close_db(_exc: Exception | None) -> None:
        db = g.pop('db', None)
        if db is not None:
            db.close()

    app.get_db = get_db  # type: ignore[attr-defined]

    with app.app_context():
        prod_db = DBAdapter(app.config['DATABASE'])
        try:
            init_db(prod_db)
            seed_default_condominio(prod_db)
            seed_default_admin(prod_db)
            seed_default_parcelas(prod_db)
            seed_default_actividades(prod_db)
            seed_default_acta_modelo(prod_db)
            cleanup_demo_environment(prod_db)
        finally:
            prod_db.close()

        if DEMO_DB_PATH.exists():
            DEMO_DB_PATH.unlink()
        demo_db = DBAdapter(str(DEMO_DB_PATH))
        try:
            init_db(demo_db)
            seed_demo_environment(demo_db)
        finally:
            demo_db.close()

    def role_required(*roles: str):
        def decorator(fn):
            @wraps(fn)
            @login_required
            def wrapper(*args, **kwargs):
                if current_user.role not in roles:
                    flash('No tienes permisos para realizar esta acción.', 'danger')
                    return redirect(url_for('dashboard'))
                return fn(*args, **kwargs)
            return wrapper
        return decorator

    def global_admin_required(fn):
        @wraps(fn)
        @login_required
        def wrapper(*args, **kwargs):
            if not current_user.is_global_admin():
                flash('Esta sección es solo para administración global.', 'danger')
                return redirect(url_for('dashboard'))
            return fn(*args, **kwargs)
        return wrapper


    LANDING_CONTACTO = {
        'telefono_principal': '+56 9 8826 2337',
        'telefono_secundario': '+56 9 8826 2337',
        'email': app.config['MAIL_TO'],
        'horario': 'Lunes a viernes · 09:00 a 18:00',
        'whatsapp_url': 'https://wa.me/56988262337',
    }

    def send_contact_request_email(nombre: str, email: str, telefono: str, condominio: str, mensaje: str) -> None:
        smtp_user = app.config['SMTP_USER']
        smtp_password = app.config['SMTP_PASSWORD']
        if not smtp_user or not smtp_password:
            raise RuntimeError('Configura SMTP_USER y SMTP_PASSWORD para enviar correos desde la landing.')

        destinatario = app.config['MAIL_TO']
        mail_from = app.config['MAIL_FROM'] or smtp_user
        asunto = f'Solicitud de información Parcelia · {nombre}'
        cuerpo = (
            'Se recibió una nueva solicitud de información desde la landing de Parcelia.\n\n'
            f'Nombre: {nombre or "-"}\n'
            f'Correo: {email or "-"}\n'
            f'Teléfono: {telefono or "-"}\n'
            f'Condominio: {condominio or "-"}\n\n'
            'Mensaje:\n'
            f'{mensaje or "-"}\n'
        )

        msg = EmailMessage()
        msg['Subject'] = asunto
        msg['From'] = mail_from
        msg['To'] = destinatario
        msg['Reply-To'] = email
        msg.set_content(cuerpo)

        smtp_host = app.config['SMTP_HOST']
        smtp_port = app.config['SMTP_PORT']

        if app.config['SMTP_USE_SSL']:
            with smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=30, context=ssl.create_default_context()) as server:
                server.login(smtp_user, smtp_password)
                server.send_message(msg)
        else:
            with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
                server.ehlo()
                if app.config['SMTP_USE_TLS']:
                    server.starttls(context=ssl.create_default_context())
                    server.ehlo()
                server.login(smtp_user, smtp_password)
                server.send_message(msg)

    @app.route('/solicitar-informacion', methods=['POST'])
    def solicitar_informacion():
        nombre = request.form.get('nombre', '').strip()
        email = request.form.get('email', '').strip()
        telefono = request.form.get('telefono', '').strip()
        condominio = request.form.get('condominio', '').strip()
        mensaje = request.form.get('mensaje', '').strip()

        if not nombre or not email or not mensaje:
            flash('Completa nombre, correo y mensaje para solicitar información.', 'danger')
            return redirect(url_for('landing'))

        app.logger.info(
            'Solicitud de información recibida | nombre=%s | email=%s | telefono=%s | condominio=%s | mensaje=%s',
            nombre,
            email,
            telefono,
            condominio,
            mensaje,
        )

        try:
            send_contact_request_email(nombre, email, telefono, condominio, mensaje)
        except Exception as exc:
            app.logger.exception('No se pudo enviar la solicitud de información por correo: %s', exc)
            flash('Recibimos tu solicitud, pero no pudimos enviarla por correo en este momento. Escríbenos a %s.' % app.config['MAIL_TO'], 'warning')
            return redirect(url_for('landing'))

        flash('Recibimos tu solicitud para Parcelia. Te contactaremos pronto.', 'success')
        return redirect(url_for('landing'))

    @app.errorhandler(400)
    def bad_request(_error):
        return render_template('error.html', error_code=400, error_title='Solicitud inválida', error_message='La solicitud no pudo procesarse. Actualiza la página e inténtalo nuevamente.'), 400

    @app.errorhandler(403)
    def forbidden(_error):
        return render_template('error.html', error_code=403, error_title='Acceso denegado', error_message='No tienes permisos para acceder a este recurso.'), 403

    @app.errorhandler(404)
    def not_found(_error):
        return render_template('error.html', error_code=404, error_title='Página no encontrada', error_message='La URL solicitada no existe o ya no está disponible.'), 404

    @app.errorhandler(413)
    def request_too_large(_error):
        return render_template('error.html', error_code=413, error_title='Archivo demasiado grande', error_message='La carga excede el tamaño máximo permitido por el servidor.'), 413

    @app.errorhandler(500)
    def internal_error(error):
        try:
            db = g.get('db')
            if db:
                db.rollback()
        except Exception:
            pass
        app.logger.exception('Error interno no controlado: %s', error)
        return render_template('error.html', error_code=500, error_title='Error interno', error_message='Ocurrió un problema inesperado. Intenta nuevamente en unos minutos.'), 500

    @app.route('/landing')
    def landing():
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))
        caracteristicas = [
            {
                'icono': '🌿',
                'titulo': 'Cobranza mensual ordenada',
                'texto': 'Genera cuotas, registra pagos y revisa morosidad por parcela sin mezclar condominios.'
            },
            {
                'icono': '🧾',
                'titulo': 'Control de gastos y mantenciones',
                'texto': 'Administra gastos operacionales, servicios y mantenciones con trazabilidad clara.'
            },
            {
                'icono': '📊',
                'titulo': 'Reportes por condominio',
                'texto': 'Visualiza ingresos, egresos, saldo y estado de pago con foco en cada comunidad.'
            },
            {
                'icono': '🏡',
                'titulo': 'Administración multi-condominio',
                'texto': 'Opera varias comunidades desde una sola plataforma con usuarios y datos separados.'
            },
            {
                'icono': '👥',
                'titulo': 'Parcelas y propietarios',
                'texto': 'Centraliza fichas, estado de cuenta y seguimiento de cada parcela o propietario.'
            },
            {
                'icono': '📝',
                'titulo': 'Actas y seguimiento',
                'texto': 'Mantén registro de acuerdos, minutas y actividades relevantes de la administración.'
            },
        ]
        return render_template('landing.html', caracteristicas=caracteristicas, landing_contacto=LANDING_CONTACTO)

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if current_user.is_authenticated:
            next_url = request.args.get('next', '').strip()
            return redirect_to_local_url(next_url, 'dashboard')
        if request.method == 'POST':
            username = request.form.get('username', '').strip().lower()
            password = request.form.get('password', '')
            is_demo_login = username in {'demo@parcelia.cl', 'comite@parcelia.cl'}
            session['db_mode'] = 'demo' if is_demo_login else 'prod'
            db = get_db()
            row = db.fetchone('SELECT u.*, c.nombre AS condominio_nombre FROM usuarios u LEFT JOIN condominios c ON c.id = u.condominio_id WHERE lower(u.username) = ? AND u.activo = 1', (username,))
            if row and check_password_hash(row['password_hash'], password):
                user = User(row, is_demo_db=is_demo_login)
                login_user(user, remember=not is_demo_login)
                session['last_activity_ts'] = int(datetime.now().timestamp())
                if user.needs_password_change():
                    flash('Debes crear una nueva contraseña para activar tu acceso.', 'warning')
                    return redirect(url_for('first_login_password'))
                flash(f'Bienvenido, {row["nombre"]}.', 'success')
                next_url = request.form.get('next', '').strip() or request.args.get('next', '').strip()
                return redirect_to_local_url(next_url, 'dashboard')
            session['db_mode'] = 'prod'
            flash('Usuario o contraseña incorrectos.', 'danger')
        demo_credentials = {'username': 'demo@parcelia.cl', 'password': '123456'}
        return render_template('login.html', demo_credentials=demo_credentials)

    @app.get('/demo')
    def demo_login():
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))
        session['db_mode'] = 'demo'
        db = get_db()
        row = db.fetchone('SELECT u.*, c.nombre AS condominio_nombre FROM usuarios u LEFT JOIN condominios c ON c.id = u.condominio_id WHERE lower(u.username) = ? AND u.activo = 1', ('demo@parcelia.cl',))
        if not row:
            session['db_mode'] = 'prod'
            flash('La cuenta demo aún no está disponible.', 'danger')
            return redirect(url_for('login'))
        user = User(row, is_demo_db=True)
        login_user(user, remember=False)
        session['last_activity_ts'] = int(datetime.now().timestamp())
        if user.needs_password_change():
            flash('Debes crear una nueva contraseña para activar tu acceso.', 'warning')
            return redirect(url_for('first_login_password'))
        flash('Entraste a la demo de Parcelia.', 'success')
        return redirect(url_for('dashboard'))

    @app.get('/logout')
    @login_required
    def logout():
        logout_user()
        session.pop('db_mode', None)
        session.pop('_csrf_token', None)
        session.pop('last_activity_ts', None)
        flash('Sesión cerrada.', 'success')
        return redirect(url_for('login'))

    @app.route('/primer-acceso', methods=['GET', 'POST'])
    @login_required
    def first_login_password():
        if not getattr(current_user, 'needs_password_change', lambda: False)():
            return redirect(url_for('dashboard'))
        if request.method == 'POST':
            current_password = request.form.get('current_password', '')
            new_password = request.form.get('new_password', '')
            confirm_password = request.form.get('confirm_password', '')

            if not check_password_hash(current_user.password_hash, current_password):
                flash('La contraseña temporal actual no es correcta.', 'danger')
            elif len(new_password) < 8:
                flash('La nueva contraseña debe tener al menos 8 caracteres.', 'danger')
            elif new_password != confirm_password:
                flash('La confirmación no coincide con la nueva contraseña.', 'danger')
            elif current_password == new_password:
                flash('La nueva contraseña debe ser distinta a la temporal.', 'danger')
            else:
                db = get_db()
                db.execute('UPDATE usuarios SET password_hash=?, must_change_password=0 WHERE id=?', (generate_password_hash(new_password), int(current_user.id.split(':')[-1]) if ':' in current_user.id else int(current_user.id)))
                db.commit()
                fresh_row = db.fetchone('SELECT u.*, c.nombre AS condominio_nombre FROM usuarios u LEFT JOIN condominios c ON c.id = u.condominio_id WHERE u.id = ?', (int(current_user.id.split(':')[-1]) if ':' in current_user.id else int(current_user.id),))
                if fresh_row:
                    login_user(User(fresh_row, is_demo_db=getattr(current_user, 'is_demo_db', False)), remember=not getattr(current_user, 'is_demo_db', False), fresh=True)
                session['last_activity_ts'] = int(datetime.now().timestamp())
                flash('Tu nueva contraseña fue guardada. Ya puedes usar el sistema.', 'success')
                return redirect(url_for('dashboard'))
        return render_template('first_login_password.html')

    @app.get('/healthz')
    def healthz():
        try:
            db = get_db()
            row = db.fetchone('SELECT 1 AS ok')
            ok = row['ok'] if row else 0
            return {'status': 'ok', 'db': int(ok)}, 200
        except Exception as exc:
            app.logger.exception('Health check failed')
            return {'status': 'error', 'detail': str(exc)}, 500

    def api_response(payload: dict[str, Any], status: int = 200, headers: dict[str, str] | None = None):
        response = app.response_class(
            response=app.json.dumps(payload),
            status=status,
            mimetype='application/json',
        )
        origin = request.headers.get('Origin', '').strip()
        allowed_origins = app.config.get('API_ALLOWED_ORIGINS', ['*'])
        if '*' in allowed_origins:
            response.headers['Access-Control-Allow-Origin'] = origin or '*'
        elif origin and origin in allowed_origins:
            response.headers['Access-Control-Allow-Origin'] = origin
        response.headers['Vary'] = 'Origin'
        response.headers['Access-Control-Allow-Headers'] = 'Authorization, Content-Type'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
        if headers:
            for k, v in headers.items():
                response.headers[k] = v
        return response

    def api_preflight_response():
        return api_response({'ok': True}, 200)

    def api_get_token_from_request() -> str:
        auth = request.headers.get('Authorization', '').strip()
        if auth.lower().startswith('bearer '):
            return auth[7:].strip()
        return request.headers.get('X-API-Token', '').strip()

    def load_api_user_from_token() -> User | None:
        token = api_get_token_from_request()
        if not token:
            return None
        try:
            payload = decode_api_token(token)
        except (SignatureExpired, BadSignature):
            return None
        mode = payload.get('mode', 'prod')
        db_url = str(DEMO_DB_PATH) if mode == 'demo' else app.config['DATABASE']
        db = DBAdapter(db_url)
        try:
            row = db.fetchone('SELECT u.*, c.nombre AS condominio_nombre FROM usuarios u LEFT JOIN condominios c ON c.id = u.condominio_id WHERE u.id = ? AND u.activo = 1', (int(payload['uid']),))
            if not row:
                return None
            return User(row, is_demo_db=(mode == 'demo'))
        finally:
            db.close()

    def api_login_required(view_func):
        @wraps(view_func)
        def wrapped(*args, **kwargs):
            user = load_api_user_from_token()
            if user is None:
                return api_response({'ok': False, 'error': 'invalid_token', 'message': 'Token ausente, inválido o expirado.'}, 401)
            g.api_user = user
            g.api_db = DBAdapter(str(DEMO_DB_PATH) if getattr(user, 'is_demo_db', False) else app.config['DATABASE'])
            try:
                return view_func(*args, **kwargs)
            finally:
                db = g.pop('api_db', None)
                if db is not None:
                    db.close()
                g.pop('api_user', None)
        return wrapped

    def api_user_to_dict(user: User) -> dict[str, Any]:
        return {
            'id': int(user.id.split(':')[-1]) if ':' in user.id else int(user.id),
            'username': user.username,
            'nombre': user.nombre,
            'role': user.role,
            'role_label': user.role_label,
            'activo': bool(user.activo),
            'condominio_id': user.condominio_id,
            'condominio_nombre': user.condominio_nombre,
            'parcela_id': user.parcela_id,
            'must_change_password': bool(getattr(user, 'must_change_password', False)),
            'demo_mode': bool(getattr(user, 'is_demo_db', False)),
        }

    def api_get_condominio_id(db: DBAdapter, user: User) -> int | None:
        selected = request.args.get('condominio_id', '').strip()
        if getattr(user, 'is_global_admin', lambda: False)() and selected.isdigit():
            return int(selected)
        if getattr(user, 'condominio_id', None):
            return int(user.condominio_id)
        return get_default_condominio_id(db)

    @app.route('/api/<path:_path>', methods=['OPTIONS'])
    def api_options(_path: str):
        return api_preflight_response()

    @app.get('/api/health')
    def api_health():
        try:
            db = get_db()
            row = db.fetchone('SELECT 1 AS ok')
            return api_response({'ok': True, 'status': 'ok', 'db': int(row['ok'] if row else 0), 'version': '1.0'})
        except Exception as exc:
            app.logger.exception('API health check failed')
            return api_response({'ok': False, 'error': 'health_error', 'message': str(exc)}, 500)

    @app.post('/api/auth/login')
    def api_auth_login():
        payload = request.get_json(silent=True) or {}
        username = str(payload.get('username', '')).strip().lower()
        password = str(payload.get('password', ''))
        mode = str(payload.get('mode', '')).strip().lower()
        is_demo_login = mode == 'demo' or username in {'demo@parcelia.cl', 'comite@parcelia.cl'}
        db = DBAdapter(str(DEMO_DB_PATH) if is_demo_login else app.config['DATABASE'])
        try:
            row = db.fetchone('SELECT u.*, c.nombre AS condominio_nombre FROM usuarios u LEFT JOIN condominios c ON c.id = u.condominio_id WHERE lower(u.username) = ? AND u.activo = 1', (username,))
            if not row or not check_password_hash(row['password_hash'], password):
                return api_response({'ok': False, 'error': 'invalid_credentials', 'message': 'Usuario o contraseña incorrectos.'}, 401)
            user = User(row, is_demo_db=is_demo_login)
            return api_response({
                'ok': True,
                'token': generate_api_token(user),
                'token_type': 'Bearer',
                'expires_in': int(app.config.get('API_TOKEN_MAX_AGE_SECONDS', 0) or 0),
                'user': api_user_to_dict(user),
                'must_change_password': bool(user.needs_password_change()),
            })
        finally:
            db.close()

    @app.get('/api/me')
    @api_login_required
    def api_me():
        return api_response({'ok': True, 'user': api_user_to_dict(g.api_user)})

    @app.post('/api/auth/change-password')
    @api_login_required
    def api_change_password():
        payload = request.get_json(silent=True) or {}
        current_password = str(payload.get('current_password', ''))
        new_password = str(payload.get('new_password', ''))
        confirm_password = str(payload.get('confirm_password', ''))
        user = g.api_user
        db = g.api_db
        if not check_password_hash(user.password_hash, current_password):
            return api_response({'ok': False, 'error': 'invalid_current_password', 'message': 'La contraseña actual no es correcta.'}, 400)
        if len(new_password) < 8:
            return api_response({'ok': False, 'error': 'weak_password', 'message': 'La nueva contraseña debe tener al menos 8 caracteres.'}, 400)
        if new_password != confirm_password:
            return api_response({'ok': False, 'error': 'password_mismatch', 'message': 'La confirmación no coincide.'}, 400)
        if current_password == new_password:
            return api_response({'ok': False, 'error': 'password_reused', 'message': 'La nueva contraseña debe ser distinta a la actual.'}, 400)
        user_id = int(user.id.split(':')[-1]) if ':' in user.id else int(user.id)
        db.execute('UPDATE usuarios SET password_hash = ?, must_change_password = 0 WHERE id = ?', (generate_password_hash(new_password), user_id))
        db.commit()
        fresh_row = db.fetchone('SELECT u.*, c.nombre AS condominio_nombre FROM usuarios u LEFT JOIN condominios c ON c.id = u.condominio_id WHERE u.id = ?', (user_id,))
        fresh_user = User(fresh_row, is_demo_db=getattr(user, 'is_demo_db', False)) if fresh_row else user
        return api_response({'ok': True, 'message': 'Contraseña actualizada correctamente.', 'user': api_user_to_dict(fresh_user), 'token': generate_api_token(fresh_user)})

    @app.get('/api/dashboard')
    @api_login_required
    def api_dashboard():
        db = g.api_db
        user = g.api_user
        condominio_id = api_get_condominio_id(db, user)
        current_month = datetime.now().strftime('%Y-%m')
        if db.kind == 'sqlite':
            resumen_mes = db.fetchone(
                """
                SELECT
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto END), 0) AS ingresos_mes,
                    COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto END), 0) AS gastos_mes,
                    COUNT(*) AS movimientos_mes
                FROM movimientos
                WHERE strftime('%Y-%m', fecha) = ? AND condominio_id = ?
                """,
                (current_month, condominio_id),
            )
        else:
            resumen_mes = db.fetchone(
                """
                SELECT
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto END), 0) AS ingresos_mes,
                    COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto END), 0) AS gastos_mes,
                    COUNT(*) AS movimientos_mes
                FROM movimientos
                WHERE to_char(fecha, 'YYYY-MM') = ? AND condominio_id = ?
                """,
                (current_month, condominio_id),
            )
        resumen_ingresos = db.fetchone('SELECT COALESCE(SUM(monto), 0) AS total FROM movimientos WHERE tipo = ? AND condominio_id = ?', ('ingreso', condominio_id))
        resumen_gastos = db.fetchone('SELECT COALESCE(SUM(monto), 0) AS total FROM movimientos WHERE tipo = ? AND condominio_id = ?', ('gasto', condominio_id))
        parcelas_activas = db.fetchone('SELECT COUNT(*) AS total FROM parcelas WHERE activo = 1 AND condominio_id = ?', (condominio_id,))
        condominio = db.fetchone('SELECT id, nombre, activo FROM condominios WHERE id = ?', (condominio_id,))
        return api_response({
            'ok': True,
            'condominio': dict(condominio) if condominio else None,
            'summary': {
                'ingresos_total': float((resumen_ingresos['total'] if resumen_ingresos else 0) or 0),
                'gastos_total': float((resumen_gastos['total'] if resumen_gastos else 0) or 0),
                'saldo_total': float(((resumen_ingresos['total'] if resumen_ingresos else 0) or 0) - ((resumen_gastos['total'] if resumen_gastos else 0) or 0)),
                'ingresos_mes': float((resumen_mes['ingresos_mes'] if resumen_mes else 0) or 0),
                'gastos_mes': float((resumen_mes['gastos_mes'] if resumen_mes else 0) or 0),
                'movimientos_mes': int((resumen_mes['movimientos_mes'] if resumen_mes else 0) or 0),
                'parcelas_activas': int((parcelas_activas['total'] if parcelas_activas else 0) or 0),
            },
        })

    @app.get('/api/parcelas')
    @api_login_required
    def api_parcelas():
        db = g.api_db
        user = g.api_user
        condominio_id = api_get_condominio_id(db, user)
        q = request.args.get('q', '').strip()
        params: list[Any] = [condominio_id]
        sql = 'SELECT id, nombre, curso, cuota_mensual, apoderado, telefono, direccion, observacion_ficha, activo, condominio_id FROM parcelas WHERE condominio_id = ?'
        if getattr(user, 'parcela_id', None):
            sql += ' AND id = ?'
            params.append(int(user.parcela_id))
        if q:
            sql += ' AND (lower(nombre) LIKE ? OR lower(COALESCE(curso, "")) LIKE ?)'
            params.extend([f'%{q.lower()}%', f'%{q.lower()}%'])
        sql += ' ORDER BY nombre'
        rows = db.fetchall(sql, params)
        return api_response({'ok': True, 'items': [dict(r) for r in rows], 'count': len(rows)})

    @app.get('/api/movimientos')
    @api_login_required
    def api_movimientos():
        db = g.api_db
        user = g.api_user
        condominio_id = api_get_condominio_id(db, user)
        try:
            limit = int(request.args.get('limit', '50') or 50)
        except Exception:
            limit = 50
        limit = min(max(limit, 1), 200)
        tipo = request.args.get('tipo', 'Todos').strip()
        q = request.args.get('q', '').strip()
        fecha_desde = request.args.get('fecha_desde', '').strip()
        fecha_hasta = request.args.get('fecha_hasta', '').strip()
        params: list[Any] = [condominio_id]
        sql = """
            SELECT m.id, m.fecha, m.tipo, m.concepto, m.monto, m.observacion, m.origen,
                   COALESCE(a.nombre, '-') AS actividad,
                   COALESCE(p.nombre, '-') AS parcela
            FROM movimientos m
            LEFT JOIN actividades a ON a.id = m.actividad_id
            LEFT JOIN parcelas p ON p.id = m.parcela_id
            WHERE m.condominio_id = ?
        """
        if getattr(user, 'parcela_id', None):
            sql += ' AND m.parcela_id = ?'
            params.append(int(user.parcela_id))
        if tipo and tipo != 'Todos':
            sql += ' AND m.tipo = ?'
            params.append(tipo)
        if q:
            sql += ' AND (lower(m.concepto) LIKE ? OR lower(COALESCE(m.observacion, "")) LIKE ?)'
            params.extend([f'%{q.lower()}%', f'%{q.lower()}%'])
        if fecha_desde:
            sql += ' AND m.fecha >= ?'
            params.append(fecha_desde)
        if fecha_hasta:
            sql += ' AND m.fecha <= ?'
            params.append(fecha_hasta)
        sql += ' ORDER BY m.fecha DESC, m.id DESC LIMIT ?'
        params.append(limit)
        rows = db.fetchall(sql, params)
        items = []
        for r in rows:
            item = dict(r)
            try:
                item['monto'] = float(item.get('monto') or 0)
            except Exception:
                pass
            items.append(item)
        return api_response({'ok': True, 'items': items, 'count': len(items)})

    @app.route('/')
    def index():
        return redirect(url_for('dashboard' if current_user.is_authenticated else 'landing'))

    @app.route('/dashboard')
    @login_required
    def dashboard():
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        resumen = db.fetchone(
            """
            SELECT
                COALESCE(SUM(CASE WHEN tipo='ingreso' THEN monto ELSE 0 END),0) ingresos,
                COALESCE(SUM(CASE WHEN tipo='gasto' THEN monto ELSE 0 END),0) gastos,
                COUNT(*) cantidad
            FROM movimientos
            WHERE condominio_id = ?
            """,
            (condominio_id,),
        )
        reporte = db.fetchall(
            """
            SELECT substr(fecha, 1, 7) AS mes,
                   COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) AS ingresos,
                   COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) AS gastos
            FROM movimientos
            WHERE condominio_id = ?
            GROUP BY substr(fecha, 1, 7)
            ORDER BY mes ASC
            """,
            (condominio_id,),
        )
        mes = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        alertas = obtener_alertas_morosidad(db, mes)
        ultimos = db.fetchall(
            """
            SELECT m.id, m.fecha, m.tipo, m.concepto, m.monto, COALESCE(a.nombre, '-') AS actividad
            FROM movimientos m
            LEFT JOIN actividades a ON a.id = m.actividad_id
            WHERE m.condominio_id = ?
            ORDER BY m.fecha DESC, m.id DESC
            LIMIT 8
            """,
            (condominio_id,),
        )
        backups = listar_backups()[:5]

        resumen_mes = db.fetchone(
            """
            SELECT
                COALESCE(SUM(CASE WHEN tipo='ingreso' THEN monto ELSE 0 END),0) ingresos_mes,
                COALESCE(SUM(CASE WHEN tipo='gasto' THEN monto ELSE 0 END),0) gastos_mes,
                COUNT(*) movimientos_mes
            FROM movimientos
            WHERE substr(fecha, 1, 7) = ? AND condominio_id = ?
            """
            ,(mes, condominio_id)
        )
        parcelas_activos = db.fetchone('SELECT COUNT(*) AS total FROM parcelas WHERE activo = 1 AND condominio_id = ?', (condominio_id,))
        cuotas = resumen_cuotas_por_parcela(db, mes)
        total_esperado = sum(float(f['cuota_mensual']) for f in cuotas if f['activo'])
        total_pagado = sum(float(f['pagado']) for f in cuotas if f['activo'])
        deuda_total = sum(max(float(f['cuota_mensual']) - float(f['pagado']), 0) for f in cuotas if f['activo'])
        parcelas_pagados = 0
        parcelas_parciales = 0
        parcelas_deuda = 0
        for fila in cuotas:
            if not fila['activo']:
                continue
            estado, _icono = estado_cuota(fila['cuota_mensual'], fila['pagado'])
            if estado == 'Pagado':
                parcelas_pagados += 1
            elif estado == 'Parcial':
                parcelas_parciales += 1
            else:
                parcelas_deuda += 1

        ingresos_mes = float(resumen_mes['ingresos_mes'] or 0)
        gastos_mes = float(resumen_mes['gastos_mes'] or 0)
        balance_total = float(resumen['ingresos'] or 0) - float(resumen['gastos'] or 0)
        balance_mes = ingresos_mes - gastos_mes
        cumplimiento = round((total_pagado / total_esperado) * 100, 1) if total_esperado else 100.0
        ultimo_mes = dict(reporte[-1]) if reporte else {'mes': mes, 'ingresos': 0, 'gastos': 0}
        ciclo_mes = obtener_ciclo_cobranza(db, mes)
        quick_actions = [
            {'label': 'Nuevo pago', 'href': url_for('pagos_new'), 'icon': '💳', 'hint': 'Registrar gasto común o abono extraordinario'},
            {'label': 'Nuevo ingreso', 'href': url_for('movimientos_new') + '?tipo=ingreso', 'icon': '➕', 'hint': 'Agregar ingreso manual'},
            {'label': 'Nuevo gasto', 'href': url_for('movimientos_new') + '?tipo=gasto', 'icon': '🧾', 'hint': 'Registrar egreso'},
            {'label': 'Reporte mensual', 'href': url_for('reportes_mensual', mes=mes), 'icon': '📊', 'hint': 'Resumen ejecutivo del mes'},
        ]

        dashboard_stats = {
            'balance_total': balance_total,
            'balance_mes': balance_mes,
            'deuda_total': deuda_total,
            'parcelas_activos': int(parcelas_activos['total'] or 0),
            'parcelas_pagados': parcelas_pagados,
            'parcelas_parciales': parcelas_parciales,
            'parcelas_deuda': parcelas_deuda,
            'ingresos_mes': ingresos_mes,
            'gastos_mes': gastos_mes,
            'movimientos_mes': int(resumen_mes['movimientos_mes'] or 0),
            'cumplimiento': cumplimiento,
            'ultimo_mes': ultimo_mes,
            'total_esperado': total_esperado,
            'total_pagado': total_pagado,
            'ciclo_generado': bool(ciclo_mes),
            'ciclo_generado_por': ciclo_mes['generado_por'] if ciclo_mes else '',
            'ciclo_generado_en': ciclo_mes['generado_en'] if ciclo_mes else '',
            'cuota_referencia': ciclo_mes['cuota_referencia'] if ciclo_mes else (cuotas[0]['cuota_mensual'] if cuotas else 40000),
        }

        return render_template(
            'dashboard.html',
            resumen=resumen,
            reporte=reporte,
            mes=mes,
            alertas=alertas,
            ultimos=ultimos,
            backups=backups,
            dashboard_stats=dashboard_stats,
            quick_actions=quick_actions,
            ciclo_mes=ciclo_mes,
        )

    @app.route('/movimientos/export/<fmt>')
    @login_required
    def movimientos_export(fmt: str):
        db = get_db()
        tipo = request.args.get('tipo', 'Todos')
        mes = request.args.get('mes', '')
        q = request.args.get('q', '').strip()
        movimientos = obtener_movimientos_filtrados(db, tipo=tipo, mes=mes, q=q)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre = f'movimientos_{ts}'
        if fmt == 'csv':
            sio = StringIO()
            writer = csv.writer(sio)
            writer.writerow(['ID', 'Fecha', 'Tipo', 'Concepto', 'Categoría', 'Parcela', 'Origen', 'Monto', 'Observación'])
            for row in movimientos:
                writer.writerow([row['id'], row['fecha'], row['tipo'], row['concepto'], row['actividad'], (row['parcela'] if ('parcela' in row.keys()) else '-'), row['origen'], row['monto'], row['observacion']])
            data = BytesIO(sio.getvalue().encode('utf-8-sig'))
            return send_file(data, mimetype='text/csv', as_attachment=True, download_name=f'{nombre}.csv')
        if fmt == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Movimientos'
            ws.append(['ID', 'Fecha', 'Tipo', 'Concepto', 'Categoría', 'Parcela', 'Origen', 'Monto', 'Observación'])
            for row in movimientos:
                ws.append([row['id'], row['fecha'], row['tipo'], row['concepto'], row['actividad'], (row['parcela'] if ('parcela' in row.keys()) else '-'), row['origen'], float(row['monto']), row['observacion']])
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
            for column in ['A','B','C','D','E','F','G','H']:
                ws.column_dimensions[column].width = 18 if column != 'D' else 34
            data = BytesIO()
            wb.save(data)
            data.seek(0)
            return send_file(data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'{nombre}.xlsx')
        if fmt == 'pdf':
            data = exportar_movimientos_pdf(movimientos, SCHOOL_NAME, SCHOOL_LOCATION, {'Tipo': tipo, 'Mes': mes or 'Todos', 'Búsqueda': q or 'Todos'})
            return send_file(data, mimetype='application/pdf', as_attachment=True, download_name=f'{nombre}.pdf')
        flash('Formato de exportación no soportado.', 'danger')
        return redirect(url_for('movimientos_list', tipo=tipo, mes=mes, q=q))

    @app.route('/backups')
    @global_admin_required
    def backups_list():
        backups = listar_backups()
        return render_template('backups.html', backups=backups)

    @app.post('/backups/crear')
    @global_admin_required
    def backups_create():
        try:
            ruta = crear_backup_db(app.config['DATABASE'])
            flash(f'Respaldo creado: {ruta.name}', 'success')
        except Exception as exc:
            flash(f'No se pudo crear el respaldo: {exc}', 'danger')
        return redirect_back('backups_list')

    @app.get('/backups/<path:nombre>')
    @global_admin_required
    def backups_download(nombre: str):
        ruta = BACKUP_DIR / nombre
        if not ruta.exists() or ruta.parent != BACKUP_DIR:
            flash('Respaldo no encontrado.', 'danger')
            return redirect(url_for('backups_list'))
        return send_file(ruta, as_attachment=True, download_name=ruta.name)

    @app.route('/condominios')
    @global_admin_required
    def condominios_list():
        db = get_db()
        condominios = db.fetchall(
            """
            SELECT c.id, c.nombre, COALESCE(c.direccion, '') AS direccion, c.activo,
                   COUNT(DISTINCT u.id) AS usuarios,
                   COUNT(DISTINCT a.id) AS parcelas
            FROM condominios c
            LEFT JOIN usuarios u ON u.condominio_id = c.id
            LEFT JOIN parcelas a ON a.condominio_id = c.id
            GROUP BY c.id, c.nombre, c.direccion, c.activo
            ORDER BY c.nombre
            """
        )
        return render_template('condominios_list.html', condominios=condominios)

    @app.route('/condominios/nuevo', methods=['GET', 'POST'])
    @global_admin_required
    def condominios_new():
        db = get_db()
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            direccion = request.form.get('direccion', '').strip()
            activo = 1 if request.form.get('activo') == 'on' else 0
            if not nombre:
                flash('El nombre del condominio es obligatorio.', 'danger')
            elif db.fetchone('SELECT 1 FROM condominios WHERE lower(nombre)=lower(?)', (nombre,)):
                flash('Ya existe un condominio con ese nombre.', 'danger')
            else:
                db.execute('INSERT INTO condominios (nombre, direccion, activo) VALUES (?, ?, ?)', (nombre, direccion, activo))
                db.commit()
                flash('Condominio creado.', 'success')
                return redirect(url_for('condominios_list'))
        return render_template('condominios_form.html', condominio=None)

    @app.route('/condominios/<int:condominio_id>/editar', methods=['GET', 'POST'])
    @global_admin_required
    def condominios_edit(condominio_id: int):
        db = get_db()
        condominio = db.fetchone('SELECT * FROM condominios WHERE id = ?', (condominio_id,))
        if not condominio:
            flash('Condominio no encontrado.', 'danger')
            return redirect(url_for('condominios_list'))
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            direccion = request.form.get('direccion', '').strip()
            activo = 1 if request.form.get('activo') == 'on' else 0
            if not nombre:
                flash('El nombre del condominio es obligatorio.', 'danger')
            else:
                db.execute('UPDATE condominios SET nombre=?, direccion=?, activo=? WHERE id=?', (nombre, direccion, activo, condominio_id))
                db.commit()
                flash('Condominio actualizado.', 'success')
                return redirect(url_for('condominios_list'))
        return render_template('condominios_form.html', condominio=condominio)

    @app.route('/usuarios')
    @role_required('admin')
    def usuarios_list():
        db = get_db()
        if current_user.is_global_admin():
            usuarios = db.fetchall("SELECT u.id, u.username, u.nombre, u.role, u.activo, u.must_change_password, c.nombre AS condominio_nombre, u.condominio_id, u.parcela_id, p.nombre AS parcela_nombre FROM usuarios u LEFT JOIN condominios c ON c.id = u.condominio_id LEFT JOIN parcelas p ON p.id = u.parcela_id ORDER BY COALESCE(c.nombre, 'ZZZ'), u.nombre, u.username")
        else:
            usuarios = db.fetchall("SELECT u.id, u.username, u.nombre, u.role, u.activo, u.must_change_password, c.nombre AS condominio_nombre, u.condominio_id, u.parcela_id, p.nombre AS parcela_nombre FROM usuarios u LEFT JOIN condominios c ON c.id = u.condominio_id LEFT JOIN parcelas p ON p.id = u.parcela_id WHERE u.condominio_id = ? ORDER BY u.nombre, u.username", (current_user.condominio_id,))
        return render_template('usuarios_list.html', usuarios=usuarios)

    @app.route('/usuarios/nuevo', methods=['GET', 'POST'])
    @role_required('admin')
    def usuarios_new():
        db = get_db()
        condominios = db.fetchall('SELECT id, nombre, activo FROM condominios WHERE activo = 1 ORDER BY nombre')
        condominio_ref = int(current_user.condominio_id) if not current_user.is_global_admin() and current_user.condominio_id else None
        parcelas_disponibles = db.fetchall('SELECT id, nombre, curso, condominio_id FROM parcelas WHERE activo = 1 ORDER BY nombre') if current_user.is_global_admin() else db.fetchall('SELECT id, nombre, curso, condominio_id FROM parcelas WHERE activo = 1 AND condominio_id = ? ORDER BY nombre', (condominio_ref,))
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            username = request.form.get('username', '').strip().lower()
            password = request.form.get('password', '')
            role = request.form.get('role', 'comite')
            activo = 1 if request.form.get('activo') == 'on' else 0
            must_change_password = 1 if request.form.get('must_change_password') else 0
            condominio_raw = request.form.get('condominio_id', '').strip()
            condominio_id = int(condominio_raw) if condominio_raw.isdigit() else None
            parcela_raw = request.form.get('parcela_id', '').strip()
            parcela_id = int(parcela_raw) if parcela_raw.isdigit() else None
            if not current_user.is_global_admin():
                condominio_id = int(current_user.condominio_id) if current_user.condominio_id else None
                if role == 'admin' and condominio_id is None:
                    flash('Debes asignar un condominio para este administrador.', 'danger')
                    return render_template('usuarios_form.html', usuario=None, roles=ALLOWED_ROLES, condominios=condominios, parcelas=parcelas_disponibles)
            if not nombre or not username or not password:
                flash('Nombre, usuario y contraseña son obligatorios.', 'danger')
            elif role not in ALLOWED_ROLES:
                flash('Rol inválido.', 'danger')
            elif not current_user.is_global_admin() and role == 'admin' and condominio_id != current_user.condominio_id:
                flash('Solo puedes crear administradores dentro de tu propio condominio.', 'danger')
            elif role != 'admin' and condominio_id is None:
                flash('Debes asignar un condominio al usuario.', 'danger')
            elif parcela_id and not db.fetchone('SELECT 1 FROM parcelas WHERE id = ? AND condominio_id = ?', (parcela_id, condominio_id)):
                flash('La parcela seleccionada no pertenece al condominio del usuario.', 'danger')
            elif parcela_id and db.fetchone('SELECT 1 FROM usuarios WHERE parcela_id = ?', (parcela_id,)):
                flash('Esa parcela ya está vinculada a otro usuario.', 'danger')
            elif db.fetchone('SELECT 1 FROM usuarios WHERE lower(username)=?', (username,)):
                flash('Ese nombre de usuario ya existe.', 'danger')
            else:
                db.execute(
                    'INSERT INTO usuarios (username, password_hash, role, nombre, activo, condominio_id, parcela_id, must_change_password) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                    (username, generate_password_hash(password), role, nombre, activo, condominio_id, parcela_id, must_change_password)
                )
                db.commit()
                flash('Usuario creado.', 'success')
                return redirect(url_for('usuarios_list'))
        return render_template('usuarios_form.html', usuario=None, roles=ALLOWED_ROLES, condominios=condominios, parcelas=parcelas_disponibles)

    @app.route('/usuarios/<int:user_id>/editar', methods=['GET', 'POST'])
    @role_required('admin')
    def usuarios_edit(user_id: int):
        db = get_db()
        usuario = db.fetchone('SELECT id, username, role, nombre, activo, condominio_id, parcela_id, must_change_password FROM usuarios WHERE id=?', (user_id,))
        if not usuario:
            flash('Usuario no encontrado.', 'danger')
            return redirect(url_for('usuarios_list'))
        if not current_user.is_global_admin() and usuario['condominio_id'] != current_user.condominio_id:
            flash('Solo puedes editar usuarios de tu propio condominio.', 'danger')
            return redirect(url_for('usuarios_list'))
        condominios = db.fetchall('SELECT id, nombre, activo FROM condominios WHERE activo = 1 ORDER BY nombre')
        parcelas_disponibles = db.fetchall('SELECT id, nombre, curso, condominio_id FROM parcelas WHERE activo = 1 ORDER BY nombre') if current_user.is_global_admin() else db.fetchall('SELECT id, nombre, curso, condominio_id FROM parcelas WHERE activo = 1 AND condominio_id = ? ORDER BY nombre', (current_user.condominio_id,))
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            username = request.form.get('username', '').strip().lower()
            password = request.form.get('password', '')
            role = request.form.get('role', 'comite')
            activo = 1 if request.form.get('activo') == 'on' else 0
            must_change_password = 1 if request.form.get('must_change_password') else 0
            condominio_raw = request.form.get('condominio_id', '').strip()
            condominio_id = int(condominio_raw) if condominio_raw.isdigit() else None
            parcela_raw = request.form.get('parcela_id', '').strip()
            parcela_id = int(parcela_raw) if parcela_raw.isdigit() else None
            if not current_user.is_global_admin():
                condominio_id = int(current_user.condominio_id) if current_user.condominio_id else None
            if not nombre or not username:
                flash('Nombre y usuario son obligatorios.', 'danger')
            elif role not in ALLOWED_ROLES:
                flash('Rol inválido.', 'danger')
            elif not current_user.is_global_admin() and role == 'admin' and condominio_id != current_user.condominio_id:
                flash('Solo puedes administrar usuarios de tu propio condominio.', 'danger')
            elif role != 'admin' and condominio_id is None:
                flash('Debes asignar un condominio al usuario.', 'danger')
            elif parcela_id and not db.fetchone('SELECT 1 FROM parcelas WHERE id = ? AND condominio_id = ?', (parcela_id, condominio_id)):
                flash('La parcela seleccionada no pertenece al condominio del usuario.', 'danger')
            elif parcela_id and db.fetchone('SELECT 1 FROM usuarios WHERE parcela_id = ? AND id <> ?', (parcela_id, user_id)):
                flash('Esa parcela ya está vinculada a otro usuario.', 'danger')
            elif db.fetchone('SELECT 1 FROM usuarios WHERE lower(username)=? AND id<>?', (username, user_id)):
                flash('Ese nombre de usuario ya existe.', 'danger')
            else:
                if password:
                    db.execute('UPDATE usuarios SET nombre=?, username=?, role=?, activo=?, condominio_id=?, parcela_id=?, password_hash=?, must_change_password=? WHERE id=?',
                               (nombre, username, role, activo, condominio_id, parcela_id, generate_password_hash(password), must_change_password, user_id))
                else:
                    db.execute('UPDATE usuarios SET nombre=?, username=?, role=?, activo=?, condominio_id=?, parcela_id=?, must_change_password=? WHERE id=?',
                               (nombre, username, role, activo, condominio_id, parcela_id, must_change_password, user_id))
                db.commit()
                flash('Usuario actualizado.', 'success')
                return redirect(url_for('usuarios_list'))
        return render_template('usuarios_form.html', usuario=usuario, roles=ALLOWED_ROLES, condominios=condominios, parcelas=parcelas_disponibles)

    @app.post('/usuarios/<int:user_id>/eliminar')
    @role_required('admin')
    def usuarios_delete(user_id: int):
        if int(current_user.id) == user_id:
            flash('No puedes eliminar tu propio usuario.', 'danger')
            return redirect(url_for('usuarios_list'))
        db = get_db()
        usuario = db.fetchone('SELECT id, condominio_id FROM usuarios WHERE id=?', (user_id,))
        if not usuario:
            flash('Usuario no encontrado.', 'danger')
            return redirect(url_for('usuarios_list'))
        if not current_user.is_global_admin() and usuario['condominio_id'] != current_user.condominio_id:
            flash('Solo puedes eliminar usuarios de tu propio condominio.', 'danger')
            return redirect(url_for('usuarios_list'))
        db.execute('DELETE FROM usuarios WHERE id=?', (user_id,))
        db.commit()
        flash('Usuario eliminado.', 'success')
        return redirect(url_for('usuarios_list'))

    @app.route('/actas')
    @login_required
    def actas_list():
        db = get_db()
        q = request.args.get('q', '').strip()
        sql = """
            SELECT id, titulo, fecha, lugar, hora_inicio, estado, asistentes, updated_at
            FROM actas
            WHERE condominio_id = ?
        """
        params: list[Any] = [get_current_condominio_id(db)]
        if q:
            sql += " AND (LOWER(COALESCE(titulo, '')) LIKE ? OR LOWER(COALESCE(lugar, '')) LIKE ? OR LOWER(COALESCE(temas, '')) LIKE ?)"
            like = sql_like_ci(q)
            params.extend([like, like, like])
        sql += " ORDER BY fecha DESC, id DESC"
        actas = db.fetchall(sql, params)
        return render_template('actas_list.html', actas=actas, q=q, ACTA_ESTADOS=ACTA_ESTADOS)
    @app.route('/actas/nueva', methods=['GET', 'POST'])
    @role_required('admin', 'presidente', 'secretario')
    def actas_new():
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        if request.method == 'POST':
            payload = extraer_acta_form(request, db)
            db.execute(
                """
                INSERT INTO actas (titulo, fecha, lugar, hora_inicio, hora_termino, asistentes, temas, desarrollo, acuerdos, responsables, observaciones, estado, created_by, updated_at, condominio_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (payload['titulo'], payload['fecha'], payload['lugar'], payload['hora_inicio'], payload['hora_termino'], payload['asistentes'], payload['temas'], payload['desarrollo'], payload['acuerdos'], payload['responsables'], payload['observaciones'], payload['estado'], current_user.nombre, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), condominio_id),
            )
            db.commit()
            flash('Acta guardada correctamente.', 'success')
            return redirect(url_for('actas_list'))
        ultimo = db.fetchone('SELECT * FROM actas WHERE condominio_id = ? ORDER BY fecha DESC, id DESC LIMIT 1', (condominio_id,))
        base = dict(ultimo) if ultimo else {
            'titulo': 'Acta Asamblea de Condominio',
            'fecha': datetime.today().strftime('%Y-%m-%d'),
            'lugar': '', 'hora_inicio': '', 'hora_termino': '', 'asistentes': '', 'temas': '',
            'desarrollo': '', 'acuerdos': '', 'responsables': '',
            'observaciones': 'Nota: En caso de no remitir observaciones luego de 48 horas hábiles de su recepción, se dará por aprobada el acta.',
            'estado': 'borrador',
        }
        return render_template('actas_form.html', acta=base, estados=ACTA_ESTADOS, duplicada=False)
    @app.route('/actas/<int:acta_id>/editar', methods=['GET', 'POST'])
    @role_required('admin', 'presidente', 'secretario')
    def actas_edit(acta_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        acta = db.fetchone('SELECT * FROM actas WHERE id = ? AND condominio_id = ?', (acta_id, condominio_id))
        if not acta:
            flash('Acta no encontrada.', 'danger')
            return redirect(url_for('actas_list'))
        if request.method == 'POST':
            payload = extraer_acta_form(request, db)
            db.execute(
                """
                UPDATE actas
                SET titulo=?, fecha=?, lugar=?, hora_inicio=?, hora_termino=?, asistentes=?, temas=?, desarrollo=?, acuerdos=?, responsables=?, observaciones=?, estado=?, updated_at=?
                WHERE id=? AND condominio_id=?
                """,
                (payload['titulo'], payload['fecha'], payload['lugar'], payload['hora_inicio'], payload['hora_termino'], payload['asistentes'], payload['temas'], payload['desarrollo'], payload['acuerdos'], payload['responsables'], payload['observaciones'], payload['estado'], datetime.now().strftime('%Y-%m-%d %H:%M:%S'), acta_id, condominio_id),
            )
            db.commit()
            flash('Acta actualizada.', 'success')
            return redirect(url_for('actas_list'))
        return render_template('actas_form.html', acta=acta, estados=ACTA_ESTADOS, duplicada=False)
    @app.post('/actas/<int:acta_id>/duplicar')
    @role_required('admin', 'presidente', 'secretario')
    def actas_duplicate(acta_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        acta = db.fetchone('SELECT * FROM actas WHERE id = ? AND condominio_id = ?', (acta_id, condominio_id))
        if not acta:
            flash('Acta no encontrada.', 'danger')
            return redirect(url_for('actas_list'))
        fecha_hoy = datetime.today().strftime('%Y-%m-%d')
        titulo = f"{acta['titulo']} - borrador {fecha_hoy}"
        db.execute(
            """
            INSERT INTO actas (titulo, fecha, lugar, hora_inicio, hora_termino, asistentes, temas, desarrollo, acuerdos, responsables, observaciones, estado, created_by, updated_at, condominio_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (titulo, fecha_hoy, acta['lugar'], acta['hora_inicio'], acta['hora_termino'], acta['asistentes'], acta['temas'], acta['desarrollo'], acta['acuerdos'], acta['responsables'], acta['observaciones'], 'borrador', current_user.nombre, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), condominio_id),
        )
        db.commit()
        flash('Se generó una nueva minuta mensual a partir de la última acta.', 'success')
        return redirect(url_for('actas_list'))
    @app.get('/actas/<int:acta_id>/pdf')
    @login_required
    def actas_pdf(acta_id: int):
        db = get_db()
        acta = db.fetchone('SELECT * FROM actas WHERE id = ? AND condominio_id = ?', (acta_id, get_current_condominio_id(db)))
        if not acta:
            flash('Acta no encontrada.', 'danger')
            return redirect(url_for('actas_list'))
        data = exportar_acta_pdf(acta)
        nombre = f"acta_{acta['fecha'] or 'sin_fecha'}_{acta_id}.pdf"
        return send_file(data, mimetype='application/pdf', as_attachment=True, download_name=nombre)



    def resolver_estado_votacion(opciones: list[dict[str, Any]], total_votos: int) -> str:
        if total_votos <= 0:
            return 'cerrada'
        ordenadas = sorted(
            opciones,
            key=lambda item: (int(item['total_votos'] or 0), -int(item['orden'] or 0), -int(item['id'])),
            reverse=True,
        )
        if len(ordenadas) >= 2 and int(ordenadas[0]['total_votos'] or 0) == int(ordenadas[1]['total_votos'] or 0):
            return 'cerrada'
        ganadora = (ordenadas[0].get('texto') or '').strip().lower()
        ganadora_norm = ganadora.replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
        if ganadora_norm in ('apruebo', 'aprobado', 'aprobada', 'si', 'sí') or 'aprueb' in ganadora_norm:
            return 'aprobada'
        if ganadora_norm in ('rechazo', 'rechazada', 'rechazado', 'no') or 'rechaz' in ganadora_norm:
            return 'rechazada'
        return 'cerrada'

    def calcular_resumen_votacion(db, votacion_id: int, condominio_id: int):
        opciones = db.fetchall(
            """
            SELECT o.id, o.texto, o.orden, COUNT(v.id) AS total_votos
            FROM votacion_opciones o
            LEFT JOIN votacion_votos v ON v.opcion_id = o.id
            WHERE o.votacion_id = ? AND o.condominio_id = ?
            GROUP BY o.id
            ORDER BY o.orden, o.id
            """,
            (votacion_id, condominio_id),
        )
        total_votos = sum(int(o['total_votos'] or 0) for o in opciones)
        estado_resuelto = resolver_estado_votacion(opciones, total_votos)
        return opciones, total_votos, estado_resuelto

    @app.get('/actas/<int:acta_id>/votaciones')
    @login_required
    def votaciones_list(acta_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        acta = db.fetchone('SELECT id, titulo, fecha, estado FROM actas WHERE id = ? AND condominio_id = ?', (acta_id, condominio_id))
        if not acta:
            flash('Acta no encontrada.', 'danger')
            return redirect(url_for('actas_list'))
        votaciones = db.fetchall(
            """
            SELECT v.*, COUNT(DISTINCT vv.id) AS total_votos, COUNT(DISTINCT vo.id) AS total_opciones
            FROM votaciones v
            LEFT JOIN votacion_votos vv ON vv.votacion_id = v.id
            LEFT JOIN votacion_opciones vo ON vo.votacion_id = v.id
            WHERE v.acta_id = ? AND v.condominio_id = ?
            GROUP BY v.id
            ORDER BY v.id DESC
            """,
            (acta_id, condominio_id),
        )
        return render_template('votaciones_list.html', acta=acta, votaciones=votaciones)

    @app.route('/actas/<int:acta_id>/votaciones/nueva', methods=['GET', 'POST'])
    @role_required('admin', 'presidente', 'secretario')
    def votacion_new(acta_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        acta = db.fetchone('SELECT id, titulo, fecha FROM actas WHERE id = ? AND condominio_id = ?', (acta_id, condominio_id))
        if not acta:
            flash('Acta no encontrada.', 'danger')
            return redirect(url_for('actas_list'))
        if request.method == 'POST':
            titulo = request.form.get('titulo', '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            opciones_texto = [line.strip() for line in request.form.get('opciones', '').splitlines() if line.strip()]
            if not titulo:
                flash('Debes indicar el título de la votación.', 'danger')
            elif len(opciones_texto) < 2:
                flash('Debes ingresar al menos dos opciones, una por línea.', 'danger')
            else:
                try:
                    if db.kind == 'postgres':
                        cur = db.execute(
                            'INSERT INTO votaciones (acta_id, titulo, descripcion, estado, created_by, created_at, condominio_id) VALUES (?, ?, ?, ?, ?, ?, ?) RETURNING id',
                            (acta_id, titulo, descripcion, 'abierta', current_user.nombre, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), condominio_id),
                        )
                        row = cur.fetchone()
                        votacion_id = row['id'] if row else None
                    else:
                        cur = db.execute(
                            'INSERT INTO votaciones (acta_id, titulo, descripcion, estado, created_by, created_at, condominio_id) VALUES (?, ?, ?, ?, ?, ?, ?)',
                            (acta_id, titulo, descripcion, 'abierta', current_user.nombre, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), condominio_id),
                        )
                        votacion_id = cur.lastrowid

                    if not votacion_id:
                        raise ValueError('No se pudo obtener el id de la votación creada.')

                    for idx, opcion in enumerate(opciones_texto, start=1):
                        db.execute('INSERT INTO votacion_opciones (votacion_id, texto, orden, condominio_id) VALUES (?, ?, ?, ?)', (votacion_id, opcion, idx, condominio_id))
                    db.commit()
                    flash('Votación creada correctamente.', 'success')
                    return redirect(url_for('votaciones_list', acta_id=acta_id))
                except Exception as e:
                    db.rollback()
                    app.logger.exception('Error creando votación')
                    flash(f'Error al crear la votación: {e}', 'danger')
        borrador = {'titulo': f"Votación de {acta['titulo']}", 'descripcion': '', 'opciones': 'Apruebo\nRechazo'}
        return render_template('votacion_form.html', acta=acta, votacion=borrador)

    @app.get('/votaciones/<int:votacion_id>')
    @login_required
    def votacion_detail(votacion_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        votacion = db.fetchone(
            """
            SELECT v.*, a.titulo AS acta_titulo, a.id AS acta_id, a.fecha AS acta_fecha
            FROM votaciones v
            INNER JOIN actas a ON a.id = v.acta_id AND a.condominio_id = v.condominio_id
            WHERE v.id = ? AND v.condominio_id = ?
            """,
            (votacion_id, condominio_id),
        )
        if not votacion:
            flash('Votación no encontrada.', 'danger')
            return redirect(url_for('actas_list'))
        opciones = db.fetchall(
            """
            SELECT o.*, COUNT(v.id) AS total_votos
            FROM votacion_opciones o
            LEFT JOIN votacion_votos v ON v.opcion_id = o.id
            WHERE o.votacion_id = ? AND o.condominio_id = ?
            GROUP BY o.id
            ORDER BY o.orden, o.id
            """,
            (votacion_id, condominio_id),
        )
        voto_usuario = db.fetchone(
            """
            SELECT vv.*, vo.texto AS opcion_texto
            FROM votacion_votos vv
            INNER JOIN votacion_opciones vo ON vo.id = vv.opcion_id
            WHERE vv.votacion_id = ? AND vv.user_id = ?
            """,
            (votacion_id, int(current_user.id)),
        )
        historial = db.fetchall(
            """
            SELECT vv.created_at, u.nombre AS usuario_nombre, u.username, p.nombre AS parcela_nombre, vo.texto AS opcion_texto
            FROM votacion_votos vv
            INNER JOIN usuarios u ON u.id = vv.user_id
            LEFT JOIN parcelas p ON p.id = vv.parcela_id
            INNER JOIN votacion_opciones vo ON vo.id = vv.opcion_id
            WHERE vv.votacion_id = ? AND vv.condominio_id = ?
            ORDER BY vv.created_at DESC, vv.id DESC
            """,
            (votacion_id, condominio_id),
        )
        total_votos = sum(int(o['total_votos'] or 0) for o in opciones)
        puede_votar = bool(getattr(current_user, 'parcela_id', None)) and votacion['estado'] == 'abierta' and voto_usuario is None
        estado_resultado = resolver_estado_votacion(opciones, total_votos) if votacion['estado'] != 'abierta' else 'abierta'
        return render_template('votacion_detail.html', votacion=votacion, opciones=opciones, voto_usuario=voto_usuario, historial=historial, total_votos=total_votos, puede_votar=puede_votar, estado_resultado=estado_resultado)

    @app.post('/votaciones/<int:votacion_id>/votar')
    @login_required
    def votacion_vote(votacion_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        votacion = db.fetchone('SELECT * FROM votaciones WHERE id = ? AND condominio_id = ?', (votacion_id, condominio_id))
        if not votacion:
            flash('Votación no encontrada.', 'danger')
            return redirect(url_for('actas_list'))
        if votacion['estado'] != 'abierta':
            flash('La votación ya está cerrada.', 'warning')
            return redirect(url_for('votacion_detail', votacion_id=votacion_id))
        parcela_id = int(current_user.parcela_id) if getattr(current_user, 'parcela_id', None) else None
        if not parcela_id:
            flash('Tu usuario no tiene una parcela asignada, por lo que no puede votar.', 'danger')
            return redirect(url_for('votacion_detail', votacion_id=votacion_id))
        if db.fetchone('SELECT 1 FROM votacion_votos WHERE votacion_id = ? AND user_id = ?', (votacion_id, int(current_user.id))):
            flash('Ya registraste tu voto en esta votación.', 'warning')
            return redirect(url_for('votacion_detail', votacion_id=votacion_id))
        opcion_raw = request.form.get('opcion_id', '').strip()
        if not opcion_raw.isdigit():
            flash('Debes seleccionar una opción válida.', 'danger')
            return redirect(url_for('votacion_detail', votacion_id=votacion_id))
        opcion = db.fetchone('SELECT id FROM votacion_opciones WHERE id = ? AND votacion_id = ? AND condominio_id = ?', (int(opcion_raw), votacion_id, condominio_id))
        if not opcion:
            flash('La opción seleccionada no pertenece a esta votación.', 'danger')
            return redirect(url_for('votacion_detail', votacion_id=votacion_id))
        db.execute('INSERT INTO votacion_votos (votacion_id, opcion_id, user_id, parcela_id, created_at, condominio_id) VALUES (?, ?, ?, ?, ?, ?)', (votacion_id, int(opcion_raw), int(current_user.id), parcela_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), condominio_id))
        db.commit()
        flash('Tu voto quedó registrado.', 'success')
        return redirect(url_for('votacion_detail', votacion_id=votacion_id))

    @app.post('/votaciones/<int:votacion_id>/cerrar')
    @role_required('admin', 'presidente', 'secretario')
    def votacion_close(votacion_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        votacion = db.fetchone('SELECT id, acta_id FROM votaciones WHERE id = ? AND condominio_id = ?', (votacion_id, condominio_id))
        if not votacion:
            flash('Votación no encontrada.', 'danger')
            return redirect(url_for('actas_list'))
        _, _, estado_resuelto = calcular_resumen_votacion(db, votacion_id, condominio_id)
        db.execute('UPDATE votaciones SET estado = ?, closed_at = ? WHERE id = ? AND condominio_id = ?', (estado_resuelto, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), votacion_id, condominio_id))
        db.commit()
        if estado_resuelto == 'aprobada':
            mensaje = 'Votación cerrada como aprobada.'
        elif estado_resuelto == 'rechazada':
            mensaje = 'Votación cerrada como rechazada.'
        else:
            mensaje = 'Votación cerrada. No fue posible resolverla como aprobada o rechazada.'
        flash(mensaje, 'success')
        return redirect(url_for('votacion_detail', votacion_id=votacion_id))


    @app.route('/votaciones/<int:votacion_id>/editar', methods=['GET', 'POST'])
    @role_required('admin', 'presidente', 'secretario')
    def votacion_edit(votacion_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        votacion = db.fetchone(
            """
            SELECT v.*, a.titulo AS acta_titulo, a.fecha AS acta_fecha
            FROM votaciones v
            INNER JOIN actas a ON a.id = v.acta_id AND a.condominio_id = v.condominio_id
            WHERE v.id = ? AND v.condominio_id = ?
            """,
            (votacion_id, condominio_id),
        )
        if not votacion:
            flash('Votación no encontrada.', 'danger')
            return redirect(url_for('actas_list'))

        opciones_actuales = db.fetchall(
            'SELECT id, texto, orden FROM votacion_opciones WHERE votacion_id = ? AND condominio_id = ? ORDER BY orden, id',
            (votacion_id, condominio_id),
        )
        tiene_votos = bool(db.fetchone('SELECT 1 FROM votacion_votos WHERE votacion_id = ? AND condominio_id = ? LIMIT 1', (votacion_id, condominio_id)))

        if request.method == 'POST':
            titulo = request.form.get('titulo', '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            opciones_texto = [line.strip() for line in request.form.get('opciones', '').splitlines() if line.strip()]
            if not titulo:
                flash('Debes indicar el título de la votación.', 'danger')
            elif not tiene_votos and len(opciones_texto) < 2:
                flash('Debes ingresar al menos dos opciones, una por línea.', 'danger')
            else:
                try:
                    db.execute(
                        'UPDATE votaciones SET titulo = ?, descripcion = ? WHERE id = ? AND condominio_id = ?',
                        (titulo, descripcion, votacion_id, condominio_id),
                    )
                    if not tiene_votos:
                        db.execute('DELETE FROM votacion_opciones WHERE votacion_id = ? AND condominio_id = ?', (votacion_id, condominio_id))
                        for idx, opcion in enumerate(opciones_texto, start=1):
                            db.execute(
                                'INSERT INTO votacion_opciones (votacion_id, texto, orden, condominio_id) VALUES (?, ?, ?, ?)',
                                (votacion_id, opcion, idx, condominio_id),
                            )
                    db.commit()
                    if tiene_votos:
                        flash('Votación actualizada. Las opciones no se modificaron porque ya tiene votos registrados.', 'success')
                    else:
                        flash('Votación actualizada correctamente.', 'success')
                    return redirect(url_for('votacion_detail', votacion_id=votacion_id))
                except Exception as e:
                    db.rollback()
                    app.logger.exception('Error editando votación')
                    flash(f'Error al editar la votación: {e}', 'danger')

        votacion_form = {
            'titulo': request.form.get('titulo', votacion['titulo']) if request.method == 'POST' else votacion['titulo'],
            'descripcion': request.form.get('descripcion', votacion['descripcion'] or '') if request.method == 'POST' else (votacion['descripcion'] or ''),
            'opciones': request.form.get('opciones', '\n'.join(o['texto'] for o in opciones_actuales)) if request.method == 'POST' else '\n'.join(o['texto'] for o in opciones_actuales),
        }
        acta = {'id': votacion['acta_id'], 'titulo': votacion['acta_titulo'], 'fecha': votacion['acta_fecha']}
        return render_template('votacion_form.html', acta=acta, votacion=votacion_form, modo='editar', tiene_votos=tiene_votos, votacion_id=votacion_id)

    @app.post('/votaciones/<int:votacion_id>/eliminar')
    @role_required('admin', 'presidente', 'secretario')
    def votacion_delete(votacion_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        votacion = db.fetchone('SELECT id, acta_id, titulo FROM votaciones WHERE id = ? AND condominio_id = ?', (votacion_id, condominio_id))
        if not votacion:
            flash('Votación no encontrada.', 'danger')
            return redirect(url_for('actas_list'))
        try:
            db.execute('DELETE FROM votacion_votos WHERE votacion_id = ? AND condominio_id = ?', (votacion_id, condominio_id))
            db.execute('DELETE FROM votacion_opciones WHERE votacion_id = ? AND condominio_id = ?', (votacion_id, condominio_id))
            db.execute('DELETE FROM votaciones WHERE id = ? AND condominio_id = ?', (votacion_id, condominio_id))
            db.commit()
            flash(f"Votación '{votacion['titulo']}' eliminada correctamente.", 'success')
        except Exception as e:
            db.rollback()
            app.logger.exception('Error eliminando votación')
            flash(f'Error al eliminar la votación: {e}', 'danger')
        return redirect(url_for('votaciones_list', acta_id=votacion['acta_id']))
    @app.route('/parcelas', endpoint='parcelas_list')
    @app.route('/parcelas')
    @login_required
    def parcelas_list():
        db = get_db()
        q = request.args.get('q', '').strip()
        mes = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        sql = """
            SELECT a.id, a.nombre, a.curso, a.cuota_mensual, a.activo,
                   COALESCE(SUM(CASE WHEN p.mes = ? THEN p.monto ELSE 0 END), 0) AS pagado_mes
            FROM parcelas a
            LEFT JOIN pagos_parcelas p ON p.parcela_id = a.id AND p.condominio_id = a.condominio_id
            WHERE a.condominio_id = ?
        """
        params: list[Any] = [mes, get_current_condominio_id(db)]
        if q:
            sql += " AND (LOWER(COALESCE(a.nombre, '')) LIKE ? OR LOWER(COALESCE(a.curso, '')) LIKE ?)"
            like = sql_like_ci(q)
            params.extend([like, like])
        sql += ' GROUP BY a.id, a.nombre, a.curso, a.cuota_mensual, a.activo ORDER BY a.nombre'
        parcelas = db.fetchall(sql, params)
        deuda_total = sum(max(float(a['cuota_mensual']) - float(a['pagado_mes']), 0) for a in parcelas if a['activo'])
        return render_template('parcelas_list.html', parcelas=parcelas, q=q, mes=mes, deuda_total=deuda_total)
    @app.route('/parcelas/nuevo', methods=['GET', 'POST'], endpoint='parcelas_new')
    @app.route('/parcelas/nuevo', methods=['GET', 'POST'])
    @role_required('admin', 'tesorero')
    def parcelas_new():
        db = get_db()
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            curso = request.form.get('curso', '').strip()
            cuota = parse_float(request.form.get('cuota_mensual', '0'))
            apoderado = request.form.get('apoderado', '').strip()
            telefono = request.form.get('telefono', '').strip()
            direccion = request.form.get('direccion', '').strip()
            observacion_ficha = request.form.get('observacion_ficha', '').strip()
            activo = 1 if request.form.get('activo') == 'on' else 0
            if not nombre:
                flash('El nombre o identificador de la parcela es obligatorio.', 'danger')
            elif parcela_duplicado(db, nombre, curso):
                flash('Ya existe una parcela con ese identificador y sector.', 'danger')
            else:
                db.execute(
                    'INSERT INTO parcelas (nombre, curso, cuota_mensual, apoderado, telefono, direccion, observacion_ficha, activo, condominio_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
                    (nombre, curso, cuota, apoderado, telefono, direccion, observacion_ficha, activo, get_current_condominio_id(db)),
                )
                db.commit()
                flash('Parcela creada correctamente.', 'success')
                return redirect(url_for('parcelas_list'))
        return render_template('parcelas_form.html', parcela=None)
    @app.route('/parcelas/<int:parcela_id>/editar', methods=['GET', 'POST'], endpoint='parcelas_edit')
    @app.route('/parcelas/<int:parcela_id>/editar', methods=['GET', 'POST'])
    @role_required('admin', 'tesorero')
    def parcelas_edit(parcela_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        parcela = db.fetchone('SELECT * FROM parcelas WHERE id = ? AND condominio_id = ?', (parcela_id, condominio_id))
        if not parcela:
            flash('Parcela no encontrada.', 'danger')
            return redirect(url_for('parcelas_list'))
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            curso = request.form.get('curso', '').strip()
            cuota = parse_float(request.form.get('cuota_mensual', '0'))
            apoderado = request.form.get('apoderado', '').strip()
            telefono = request.form.get('telefono', '').strip()
            direccion = request.form.get('direccion', '').strip()
            observacion_ficha = request.form.get('observacion_ficha', '').strip()
            activo = 1 if request.form.get('activo') == 'on' else 0
            if not nombre:
                flash('El nombre o identificador de la parcela es obligatorio.', 'danger')
            elif parcela_duplicado(db, nombre, curso, exclude_id=parcela_id):
                flash('Ya existe otra parcela con ese identificador y sector.', 'danger')
            else:
                db.execute(
                    'UPDATE parcelas SET nombre = ?, curso = ?, cuota_mensual = ?, apoderado = ?, telefono = ?, direccion = ?, observacion_ficha = ?, activo = ? WHERE id = ? AND condominio_id = ?',
                    (nombre, curso, cuota, apoderado, telefono, direccion, observacion_ficha, activo, parcela_id, condominio_id),
                )
                db.commit()
                flash('Parcela actualizada.', 'success')
                return redirect(url_for('parcelas_list'))
        return render_template('parcelas_form.html', parcela=parcela)
    @app.post('/parcelas/<int:parcela_id>/eliminar', endpoint='parcelas_delete')
    @app.post('/parcelas/<int:parcela_id>/eliminar')
    @role_required('admin', 'tesorero')
    def parcelas_delete(parcela_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        parcela = db.fetchone('SELECT nombre FROM parcelas WHERE id=? AND condominio_id = ?', (parcela_id, condominio_id))
        if not parcela:
            flash('Parcela no encontrada.', 'danger')
            return redirect(url_for('parcelas_list'))
        db.execute('DELETE FROM pagos_parcelas WHERE parcela_id = ? AND condominio_id = ?', (parcela_id, condominio_id))
        db.execute('DELETE FROM parcelas WHERE id = ? AND condominio_id = ?', (parcela_id, condominio_id))
        db.commit()
        flash(f'Parcela eliminada: {parcela["nombre"]}.', 'success')
        return redirect(url_for('parcelas_list'))
    @app.route('/parcelas/<int:parcela_id>', endpoint='parcela_detail')
    @app.route('/parcelas/<int:parcela_id>')
    @login_required
    def parcela_detail(parcela_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        parcela = db.fetchone('SELECT * FROM parcelas WHERE id = ? AND condominio_id = ?', (parcela_id, condominio_id))
        if not parcela:
            flash('Parcela no encontrada.', 'danger')
            return redirect(url_for('parcelas_list'))
        mes_actual = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        resumen_mes = db.fetchone(
            """
            SELECT COALESCE(SUM(CASE WHEN p.mes = ? THEN p.monto ELSE 0 END), 0) AS pagado_mes,
                   COUNT(CASE WHEN p.mes = ? THEN 1 END) AS pagos_mes
            FROM pagos_parcelas p
            WHERE p.parcela_id = ? AND p.condominio_id = ?
            """,
            (mes_actual, mes_actual, parcela_id, condominio_id),
        )
        resumen_aportes = db.fetchone(
            """
            SELECT COALESCE(SUM(CASE WHEN m.tipo = 'ingreso' THEN m.monto ELSE 0 END), 0) AS ingresos_actividad,
                   COALESCE(SUM(CASE WHEN m.tipo = 'gasto' THEN m.monto ELSE 0 END), 0) AS gastos_asociados,
                   COUNT(*) AS movimientos_asociados
            FROM movimientos m
            WHERE m.parcela_id = ? AND m.condominio_id = ?
            """,
            (parcela_id, condominio_id),
        )
        historial_cuotas = db.fetchall(
            """
            SELECT p.id, p.fecha, p.mes, p.monto, p.observacion, 'Gasto común' AS tipo, NULL AS actividad
            FROM pagos_parcelas p
            WHERE p.parcela_id = ? AND p.condominio_id = ?
            ORDER BY p.fecha DESC, p.id DESC
            """,
            (parcela_id, condominio_id),
        )
        historial_aportes = db.fetchall(
            """
            SELECT m.id, m.fecha, substr(m.fecha,1,7) AS mes, m.monto, m.observacion,
                   CASE WHEN m.tipo = 'gasto' THEN 'Egreso asociado' ELSE 'Ingreso extraordinario' END AS tipo,
                   COALESCE(a.nombre, '-') AS actividad
            FROM movimientos m
            LEFT JOIN actividades a ON a.id = m.actividad_id
            WHERE m.parcela_id = ? AND m.condominio_id = ?
            ORDER BY m.fecha DESC, m.id DESC
            """,
            (parcela_id, condominio_id),
        )
        historial = sorted([dict(x) for x in historial_cuotas] + [dict(x) for x in historial_aportes], key=lambda x: (x['fecha'], x['id']), reverse=True)
        actividad_resumen = db.fetchall(
            """
            SELECT COALESCE(a.id, 0) AS actividad_id, COALESCE(a.nombre, 'Sin actividad') AS actividad,
                   COALESCE(SUM(CASE WHEN m.tipo = 'ingreso' THEN m.monto ELSE 0 END), 0) AS ingresos,
                   COALESCE(SUM(CASE WHEN m.tipo = 'gasto' THEN m.monto ELSE 0 END), 0) AS egresos,
                   COUNT(*) AS movimientos
            FROM movimientos m
            LEFT JOIN actividades a ON a.id = m.actividad_id
            WHERE m.parcela_id = ? AND m.condominio_id = ?
            GROUP BY a.id, a.nombre
            ORDER BY actividad
            """,
            (parcela_id, condominio_id),
        )
        deuda_mes = max(float(parcela['cuota_mensual']) - float(resumen_mes['pagado_mes'] or 0), 0) if parcela['activo'] else 0
        resumen = {
            'mes': mes_actual, 'pagado_mes': float(resumen_mes['pagado_mes'] or 0), 'pagos_mes': int(resumen_mes['pagos_mes'] or 0),
            'deuda_mes': deuda_mes, 'ingresos_actividad': float(resumen_aportes['ingresos_actividad'] or 0),
            'gastos_asociados': float(resumen_aportes['gastos_asociados'] or 0), 'movimientos_asociados': int(resumen_aportes['movimientos_asociados'] or 0),
        }
        ficha = {'apoderado': parcela['apoderado'], 'telefono': parcela['telefono'], 'direccion': parcela['direccion'], 'observacion_ficha': parcela['observacion_ficha']}
        return render_template('parcela_detail.html', parcela=parcela, historial=historial, resumen=resumen, actividad_resumen=actividad_resumen, ficha=ficha)
    # Ruta legacy deshabilitada en refactor multi-condominio
    # Ruta legacy deshabilitada en refactor multi-condominio
    def pagos_list_legacy():
        db = get_db()
        mes = request.args.get('mes', '').strip()
        sql = """
            SELECT p.id, p.parcela_id, a.nombre, a.curso, p.fecha, p.mes, p.monto, p.observacion, p.movimiento_id
            FROM pagos_parcelas p
            INNER JOIN parcelas a ON a.id = p.parcela_id
            WHERE 1=1
        """
        params: list[Any] = []
        if mes:
            sql += ' AND p.mes = ?'
            params.append(mes)
        sql += ' ORDER BY p.fecha DESC, a.nombre'
        pagos = db.fetchall(sql, params)
        return render_template('pagos_list.html', pagos=pagos, mes=mes)

    # Ruta legacy deshabilitada en refactor multi-condominio
    # Ruta legacy deshabilitada en refactor multi-condominio
    def pagos_new_legacy():
        db = get_db()
        parcelas = db.fetchall('SELECT id, nombre, curso, cuota_mensual FROM parcelas WHERE activo = 1 ORDER BY nombre')
        actividades = db.fetchall('SELECT id, nombre, fecha FROM actividades ORDER BY fecha DESC, nombre')
        if request.method == 'POST':
            parcela_id = int(request.form.get('parcela_id', '0') or 0)
            fecha = request.form.get('fecha', '').strip()
            mes = request.form.get('mes', '').strip()
            monto = parse_float(request.form.get('monto', '0'))
            observacion = request.form.get('observacion', '').strip()
            tipo_pago = request.form.get('tipo_pago', 'cuota_mensual')
            actividad_raw = request.form.get('actividad_id', '').strip()
            actividad_id = int(actividad_raw) if actividad_raw else None
            try:
                validar_fecha(fecha)
                datetime.strptime(mes + '-01', '%Y-%m-%d')
            except Exception:
                flash('Fecha o mes inválido.', 'danger')
                return render_template('pagos_form.html', parcelas=parcelas, actividades=actividades, pago=None)
            if tipo_pago == 'actividad_parcela' and not actividad_id:
                flash('Debes seleccionar una categoría para un ingreso extraordinario asociado.', 'danger')
            elif tipo_pago == 'cuota_mensual' and pago_duplicado(db, parcela_id, mes):
                flash('Esa parcela ya tiene un pago de gasto común registrado para ese mes.', 'danger')
            else:
                registrar_pago_parcela(db, parcela_id, fecha, mes, monto, observacion, actividad_id, tipo_pago)
                db.commit()
                flash('Pago registrado correctamente.', 'success')
    @app.route('/pagos')
    @login_required
    def pagos_list():
        db = get_db()
        mes = request.args.get('mes', '').strip()
        sql = """
            SELECT p.id, p.parcela_id, a.nombre, a.curso, p.fecha, p.mes, p.monto, p.observacion, p.movimiento_id
            FROM pagos_parcelas p
            INNER JOIN parcelas a ON a.id = p.parcela_id
            WHERE p.condominio_id = ?
        """
        params: list[Any] = [get_current_condominio_id(db)]
        if mes:
            sql += ' AND p.mes = ?'
            params.append(mes)
        sql += ' ORDER BY p.fecha DESC, a.nombre'
        pagos = db.fetchall(sql, params)
        return render_template('pagos_list.html', pagos=pagos, mes=mes)
    @app.route('/pagos/nuevo', methods=['GET', 'POST'])
    @role_required('admin', 'tesorero')
    def pagos_new():
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        parcelas = db.fetchall('SELECT id, nombre, curso, cuota_mensual FROM parcelas WHERE activo = 1 AND condominio_id = ? ORDER BY nombre', (condominio_id,))
        actividades = db.fetchall('SELECT id, nombre, fecha FROM actividades WHERE condominio_id = ? ORDER BY fecha DESC, nombre', (condominio_id,))
        if request.method == 'POST':
            parcela_id = int(request.form.get('parcela_id', '0') or 0)
            fecha = request.form.get('fecha', '').strip()
            mes = request.form.get('mes', '').strip()
            monto = parse_float(request.form.get('monto', '0'))
            observacion = request.form.get('observacion', '').strip()
            tipo_pago = request.form.get('tipo_pago', 'cuota_mensual')
            actividad_raw = request.form.get('actividad_id', '').strip()
            actividad_id = int(actividad_raw) if actividad_raw else None
            try:
                validar_fecha(fecha)
                datetime.strptime(mes + '-01', '%Y-%m-%d')
            except Exception:
                flash('Fecha o mes inválido.', 'danger')
                return render_template('pagos_form.html', parcelas=parcelas, actividades=actividades, pago=None)
            if tipo_pago == 'actividad_parcela' and not actividad_id:
                flash('Debes seleccionar una categoría para un ingreso extraordinario asociado.', 'danger')
            elif tipo_pago == 'cuota_mensual' and pago_duplicado(db, parcela_id, mes):
                flash('Esa parcela ya tiene un pago de gasto común registrado para ese mes.', 'danger')
            else:
                registrar_pago_parcela(db, parcela_id, fecha, mes, monto, observacion, actividad_id, tipo_pago)
                db.commit()
                flash('Pago registrado correctamente.', 'success')
                return redirect(url_for('pagos_list'))
        return render_template('pagos_form.html', parcelas=parcelas, actividades=actividades, pago=None)
    @app.route('/pagos/<int:pago_id>/editar', methods=['GET', 'POST'])
    @role_required('admin', 'tesorero')
    def pagos_edit(pago_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        pago = db.fetchone('SELECT * FROM pagos_parcelas WHERE id=? AND condominio_id = ?', (pago_id, condominio_id))
        if not pago:
            flash('Pago no encontrado.', 'danger')
            return redirect(url_for('pagos_list'))
        parcelas = db.fetchall('SELECT id, nombre, curso, cuota_mensual FROM parcelas WHERE condominio_id = ? AND (activo = 1 OR id = ?) ORDER BY nombre', (condominio_id, pago['parcela_id']))
        actividades = db.fetchall('SELECT id, nombre, fecha FROM actividades WHERE condominio_id = ? ORDER BY fecha DESC, nombre', (condominio_id,))
        if request.method == 'POST':
            parcela_id = int(request.form.get('parcela_id', '0') or 0)
            fecha = request.form.get('fecha', '').strip()
            mes = request.form.get('mes', '').strip()
            monto = parse_float(request.form.get('monto', '0'))
            observacion = request.form.get('observacion', '').strip()
            try:
                validar_fecha(fecha)
                datetime.strptime(mes + '-01', '%Y-%m-%d')
            except Exception:
                flash('Fecha o mes inválido.', 'danger')
                return render_template('pagos_form.html', parcelas=parcelas, actividades=actividades, pago=pago)
            if db.fetchone('SELECT 1 FROM pagos_parcelas WHERE parcela_id=? AND mes=? AND id<>? AND condominio_id = ?', (parcela_id, mes, pago_id, condominio_id)):
                flash('Esa parcela ya tiene otro pago registrado para ese mes.', 'danger')
            else:
                db.execute('UPDATE pagos_parcelas SET parcela_id=?, fecha=?, mes=?, monto=?, observacion=? WHERE id=? AND condominio_id = ?',
                           (parcela_id, fecha, mes, monto, observacion, pago_id, condominio_id))
                db.execute('UPDATE movimientos SET fecha=?, concepto=?, monto=?, parcela_id=?, observacion=? WHERE id=? AND condominio_id = ?',
                           (fecha, f'Gasto común parcela: {obtener_nombre_parcela(db, parcela_id)} ({mes})', monto, parcela_id, observacion, pago['movimiento_id'], condominio_id))
                db.commit()
                flash('Pago actualizado.', 'success')
                return redirect(url_for('pagos_list'))
        return render_template('pagos_form.html', parcelas=parcelas, actividades=actividades, pago=pago)
    @app.post('/pagos/<int:pago_id>/eliminar')
    @role_required('admin', 'tesorero')
    def pagos_delete(pago_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        pago = db.fetchone('SELECT * FROM pagos_parcelas WHERE id=? AND condominio_id = ?', (pago_id, condominio_id))
        if not pago:
            flash('Pago no encontrado.', 'danger')
            return redirect(url_for('pagos_list'))
        db.execute('DELETE FROM pagos_parcelas WHERE id=? AND condominio_id = ?', (pago_id, condominio_id))
        if pago['movimiento_id']:
            db.execute('DELETE FROM movimientos WHERE id=? AND condominio_id = ?', (pago['movimiento_id'], condominio_id))
        db.commit()
        flash('Pago eliminado.', 'success')
        return redirect(url_for('pagos_list'))
    @app.route('/movimientos')
    @login_required
    def movimientos_list():
        db = get_db()
        tipo = request.args.get('tipo', 'Todos')
        mes = request.args.get('mes', '')
        q = request.args.get('q', '').strip()
        fecha_desde = request.args.get('fecha_desde', '').strip()
        fecha_hasta = request.args.get('fecha_hasta', '').strip()
        actividad_id = request.args.get('actividad_id', '').strip()
        parcela_id = request.args.get('parcela_id', '').strip()
        movimientos = obtener_movimientos_filtrados(db, tipo=tipo, mes=mes, q=q, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, actividad_id=actividad_id, parcela_id=parcela_id)
        actividades = db.fetchall('SELECT id, nombre, fecha FROM actividades WHERE condominio_id = ? ORDER BY fecha DESC, nombre', (get_current_condominio_id(db),))
        parcelas = db.fetchall('SELECT id, nombre, curso FROM parcelas WHERE activo = 1 AND condominio_id = ? ORDER BY nombre', (get_current_condominio_id(db),))
        return render_template('movimientos_list.html', movimientos=movimientos, tipo=tipo, mes=mes, q=q, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, actividad_id=actividad_id, parcela_id=parcela_id, actividades=actividades, parcelas=parcelas)
    @app.route('/movimientos/nuevo', methods=['GET', 'POST'])
    @role_required('admin', 'tesorero')
    def movimientos_new():
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        actividades = db.fetchall('SELECT id, nombre, fecha FROM actividades WHERE condominio_id = ? ORDER BY fecha DESC, nombre', (condominio_id,))
        parcelas = db.fetchall('SELECT id, nombre, curso FROM parcelas WHERE activo = 1 AND condominio_id = ? ORDER BY nombre', (condominio_id,))
        if request.method == 'POST':
            fecha = request.form.get('fecha', '').strip()
            tipo = request.form.get('tipo', 'ingreso').strip()
            concepto = request.form.get('concepto', '').strip()
            monto = parse_float(request.form.get('monto', '0'))
            actividad_raw = request.form.get('actividad_id', '').strip()
            actividad_id = int(actividad_raw) if actividad_raw else None
            parcela_raw = request.form.get('parcela_id', '').strip()
            parcela_id = int(parcela_raw) if parcela_raw else None
            observacion = request.form.get('observacion', '').strip()
            try:
                validar_fecha(fecha)
            except Exception:
                flash('Fecha inválida.', 'danger')
                return render_template('movimientos_form.html', actividades=actividades, parcelas=parcelas, movimiento=None)
            db.execute(
                'INSERT INTO movimientos (fecha, tipo, concepto, monto, actividad_id, parcela_id, observacion, origen, condominio_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (fecha, tipo, concepto, monto, actividad_id, parcela_id, observacion, 'general', condominio_id),
            )
            db.commit()
            flash('Movimiento creado.', 'success')
            next_url = request.form.get('next', '').strip()
            return redirect_to_local_url(next_url, 'movimientos_list')
        return render_template('movimientos_form.html', actividades=actividades, parcelas=parcelas, movimiento=None)
    @app.route('/movimientos/<int:movimiento_id>/editar', methods=['GET', 'POST'])
    @role_required('admin', 'tesorero')
    def movimientos_edit(movimiento_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        if getattr(current_user, 'is_global_admin', lambda: False)():
            movimiento = db.fetchone('SELECT * FROM movimientos WHERE id=?', (movimiento_id,))
            if movimiento:
                condominio_id = movimiento['condominio_id']
        else:
            movimiento = db.fetchone('SELECT * FROM movimientos WHERE id=? AND condominio_id = ?', (movimiento_id, condominio_id))
        if not movimiento:
            flash('Movimiento no encontrado.', 'danger')
            return redirect(url_for('movimientos_list', condominio_id=request.values.get('condominio_id', '').strip() or None))
        actividades = db.fetchall('SELECT id, nombre, fecha FROM actividades WHERE condominio_id = ? ORDER BY fecha DESC, nombre', (condominio_id,))
        parcelas = db.fetchall('SELECT id, nombre, curso FROM parcelas WHERE activo = 1 AND condominio_id = ? ORDER BY nombre', (condominio_id,))
        if request.method == 'POST':
            fecha = request.form.get('fecha', '').strip()
            tipo = request.form.get('tipo', 'ingreso').strip()
            concepto = request.form.get('concepto', '').strip()
            monto = parse_float(request.form.get('monto', '0'))
            actividad_raw = request.form.get('actividad_id', '').strip()
            actividad_id = int(actividad_raw) if actividad_raw else None
            parcela_raw = request.form.get('parcela_id', '').strip()
            parcela_id = int(parcela_raw) if parcela_raw else None
            observacion = request.form.get('observacion', '').strip()
            try:
                validar_fecha(fecha)
            except Exception:
                flash('Fecha inválida.', 'danger')
                return render_template('movimientos_form.html', actividades=actividades, parcelas=parcelas, movimiento=movimiento)
            db.execute('UPDATE movimientos SET fecha=?, tipo=?, concepto=?, monto=?, actividad_id=?, parcela_id=?, observacion=? WHERE id=? AND condominio_id = ?',
                       (fecha, tipo, concepto, monto, actividad_id, parcela_id, observacion, movimiento_id, condominio_id))
            db.commit()
            flash('Movimiento actualizado.', 'success')
            next_url = request.form.get('next', '').strip()
            return redirect_to_local_url(next_url, 'movimientos_list')
        return render_template('movimientos_form.html', actividades=actividades, parcelas=parcelas, movimiento=movimiento)
    @app.post('/movimientos/<int:movimiento_id>/eliminar')
    @role_required('admin', 'tesorero')
    def movimientos_delete(movimiento_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        if getattr(current_user, 'is_global_admin', lambda: False)():
            movimiento = db.fetchone('SELECT concepto, condominio_id FROM movimientos WHERE id=?', (movimiento_id,))
            if movimiento:
                condominio_id = movimiento['condominio_id']
        else:
            movimiento = db.fetchone('SELECT concepto, condominio_id FROM movimientos WHERE id=? AND condominio_id = ?', (movimiento_id, condominio_id))
        if not movimiento:
            flash('Movimiento no encontrado.', 'danger')
            return redirect(url_for('movimientos_list', condominio_id=request.values.get('condominio_id', '').strip() or None))
        db.execute('DELETE FROM pagos_parcelas WHERE movimiento_id = ? AND condominio_id = ?', (movimiento_id, condominio_id))
        db.execute('DELETE FROM movimientos WHERE id = ? AND condominio_id = ?', (movimiento_id, condominio_id))
        db.commit()
        flash(f'Movimiento eliminado: {movimiento["concepto"]}.', 'success')
        next_url = request.form.get('next', '').strip()
        return redirect_to_local_url(next_url, 'movimientos_list')
    @app.route('/actividades')
    @login_required
    def actividades_list():
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        actividades = db.fetchall(
            """
            SELECT a.id, a.nombre, a.fecha, COALESCE(a.descripcion, '') AS descripcion,
                   COALESCE(SUM(CASE WHEN m.tipo = 'ingreso' THEN m.monto ELSE 0 END), 0) AS ingresos,
                   COALESCE(SUM(CASE WHEN m.tipo = 'gasto' THEN m.monto ELSE 0 END), 0) AS egresos,
                   COALESCE(SUM(CASE WHEN m.tipo = 'ingreso' THEN 1 ELSE 0 END), 0) AS cantidad_ingresos,
                   COALESCE(SUM(CASE WHEN m.tipo = 'gasto' THEN 1 ELSE 0 END), 0) AS cantidad_egresos
            FROM actividades a
            LEFT JOIN movimientos m ON m.actividad_id = a.id AND m.condominio_id = a.condominio_id
            WHERE a.condominio_id = ?
            GROUP BY a.id, a.nombre, a.fecha, a.descripcion
            ORDER BY a.fecha DESC, a.nombre
            """,
            (condominio_id,),
        )
        resumen_general = {'ingresos': sum(float(a['ingresos'] or 0) for a in actividades), 'egresos': sum(float(a['egresos'] or 0) for a in actividades), 'cantidad_actividades': len(actividades)}
        resumen_general['balance'] = resumen_general['ingresos'] - resumen_general['egresos']
        return render_template('actividades_list.html', actividades=actividades, resumen_general=resumen_general)
    @app.route('/actividades/reporte')
    @login_required
    def actividades_report():
        db = get_db()
        mes = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        condominio_id = get_current_condominio_id(db)
        actividades = db.fetchall(
            """
            SELECT a.id, a.nombre, a.fecha, COALESCE(a.descripcion, '') AS descripcion,
                   COALESCE(SUM(CASE WHEN m.tipo = 'ingreso' THEN m.monto ELSE 0 END), 0) AS ingresos,
                   COALESCE(SUM(CASE WHEN m.tipo = 'gasto' THEN m.monto ELSE 0 END), 0) AS egresos,
                   COUNT(m.id) AS movimientos
            FROM actividades a
            LEFT JOIN movimientos m ON m.actividad_id = a.id AND m.condominio_id = a.condominio_id
            WHERE a.condominio_id = ?
            GROUP BY a.id, a.nombre, a.fecha, a.descripcion
            ORDER BY a.fecha DESC, a.nombre
            """,
            (condominio_id,),
        )
        deudas = resumen_cuotas_por_parcela(db, mes)
        total_deuda = sum(max(float(f['cuota_mensual']) - float(f['pagado']), 0) for f in deudas if f['activo'])
        return render_template('actividades_report.html', actividades=actividades, mes=mes, deudas=deudas, total_deuda=total_deuda)
    @app.route('/actividades/<int:actividad_id>')
    @login_required
    def actividad_detail(actividad_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        actividad = db.fetchone('SELECT id, nombre, fecha, COALESCE(descripcion, '') AS descripcion FROM actividades WHERE id = ? AND condominio_id = ?', (actividad_id, condominio_id))
        if not actividad:
            flash('Actividad no encontrada.', 'danger')
            return redirect(url_for('actividades_list'))
        resumen = db.fetchone(
            """
            SELECT COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) AS ingresos,
                   COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN monto ELSE 0 END), 0) AS gastos,
                   COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN 1 ELSE 0 END), 0) AS cantidad_ingresos,
                   COALESCE(SUM(CASE WHEN tipo = 'gasto' THEN 1 ELSE 0 END), 0) AS cantidad_gastos
            FROM movimientos WHERE actividad_id = ? AND condominio_id = ?
            """,
            (actividad_id, condominio_id),
        )
        ingresos = db.fetchall(
            """
            SELECT m.id, m.fecha, m.concepto, m.monto, COALESCE(m.observacion, '') AS observacion, COALESCE(m.origen, 'general') AS origen, COALESCE(al.nombre, '-') AS parcela
            FROM movimientos m LEFT JOIN parcelas al ON al.id = m.parcela_id
            WHERE m.actividad_id = ? AND m.tipo = 'ingreso' AND m.condominio_id = ?
            ORDER BY m.fecha DESC, m.id DESC
            """,
            (actividad_id, condominio_id),
        )
        gastos = db.fetchall(
            """
            SELECT m.id, m.fecha, m.concepto, m.monto, COALESCE(m.observacion, '') AS observacion, COALESCE(m.origen, 'general') AS origen
            FROM movimientos m
            WHERE m.actividad_id = ? AND m.tipo = 'gasto' AND m.condominio_id = ?
            ORDER BY m.fecha DESC, m.id DESC
            """,
            (actividad_id, condominio_id),
        )
        return render_template('actividad_detail.html', actividad=actividad, resumen=resumen, ingresos=ingresos, gastos=gastos)
    @app.route('/actividades/nueva', methods=['GET', 'POST'])
    @role_required('admin', 'tesorero')
    def actividades_new():
        db = get_db()
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            fecha = request.form.get('fecha', '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            try:
                validar_fecha(fecha)
            except Exception:
                flash('Fecha inválida.', 'danger')
                return render_template('actividades_form.html', actividad=None)
            db.execute('INSERT INTO actividades (nombre, fecha, descripcion, condominio_id) VALUES (?, ?, ?, ?)', (nombre, fecha, descripcion, get_current_condominio_id(db)))
            db.commit()
            flash('Actividad creada.', 'success')
            return redirect(url_for('actividades_list'))
        return render_template('actividades_form.html', actividad=None)

    @app.route('/actividades/<int:actividad_id>/editar', methods=['GET', 'POST'])
    @role_required('admin', 'tesorero')
    def actividades_edit(actividad_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        actividad = db.fetchone('SELECT * FROM actividades WHERE id=? AND condominio_id = ?', (actividad_id, condominio_id))
        if not actividad:
            flash('Actividad no encontrada.', 'danger')
            return redirect(url_for('actividades_list'))
        if request.method == 'POST':
            nombre = request.form.get('nombre', '').strip()
            fecha = request.form.get('fecha', '').strip()
            descripcion = request.form.get('descripcion', '').strip()
            try:
                validar_fecha(fecha)
            except Exception:
                flash('Fecha inválida.', 'danger')
                return render_template('actividades_form.html', actividad=actividad)
            db.execute('UPDATE actividades SET nombre=?, fecha=?, descripcion=? WHERE id=? AND condominio_id = ?', (nombre, fecha, descripcion, actividad_id, condominio_id))
            db.commit()
            flash('Actividad actualizada.', 'success')
            return redirect(url_for('actividades_list'))
        return render_template('actividades_form.html', actividad=actividad)

    @app.post('/actividades/<int:actividad_id>/eliminar')
    @role_required('admin', 'tesorero')
    def actividades_delete(actividad_id: int):
        db = get_db()
        condominio_id = get_current_condominio_id(db)
        actividad = db.fetchone('SELECT nombre FROM actividades WHERE id=? AND condominio_id = ?', (actividad_id, condominio_id))
        if not actividad:
            flash('Actividad no encontrada.', 'danger')
            return redirect(url_for('actividades_list'))
        db.execute('UPDATE movimientos SET actividad_id = NULL WHERE actividad_id = ? AND condominio_id = ?', (actividad_id, condominio_id))
        db.execute('DELETE FROM actividades WHERE id = ? AND condominio_id = ?', (actividad_id, condominio_id))
        db.commit()
        flash(f'Actividad eliminada: {actividad["nombre"]}.', 'success')
        return redirect(url_for('actividades_list'))


    @app.route('/cuotas')
    @login_required
    def cuotas_view():
        db = get_db()
        mes = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        exportar = request.args.get('exportar', '').strip().lower()
        filtro_reporte = request.args.get('filtro_reporte', 'deuda').strip().lower()
        if filtro_reporte not in ('deuda', 'todos'):
            filtro_reporte = 'deuda'

        if exportar == 'pdf':
            pdf_buffer = construir_pdf_deudores(db, mes, filtro_reporte)
            filename = f'reporte_cuotas_{filtro_reporte}_{mes}.pdf'
            return send_file(pdf_buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

        filas = resumen_cuotas_por_parcela(db, mes)
        if filtro_reporte == 'deuda':
            filas = [
                fila for fila in filas
                if fila['activo'] and max(float(fila['cuota_mensual']) - float(fila['pagado']), 0) > 0
            ]

        total_esperado = sum(float(x['cuota_mensual']) for x in filas if x['activo'])
        total_pagado = sum(float(x['pagado']) for x in filas)
        total_debe = sum(max(float(x['cuota_mensual']) - float(x['pagado']), 0) for x in filas if x['activo'])
        alertas = obtener_alertas_morosidad(db, mes)
        if filtro_reporte == 'deuda':
            alertas = [alerta for alerta in alertas if alerta['debe'] > 0]
        return render_template(
            'cuotas.html',
            filas=filas,
            mes=mes,
            total_esperado=total_esperado,
            total_pagado=total_pagado,
            total_debe=total_debe,
            alertas=alertas,
            filtro_reporte=filtro_reporte,
        )

    @app.post('/cuotas/generar')
    @role_required('admin', 'tesorero')
    def cuotas_generar():
        db = get_db()
        mes = request.form.get('mes') or datetime.today().strftime('%Y-%m')
        try:
            datetime.strptime(mes + '-01', '%Y-%m-%d')
            ciclo = generar_ciclo_cobranza(db, mes, current_user.nombre)
            db.commit()
            if ciclo['creado']:
                flash(f'Ciclo de cobranza {mes} generado con cuota de referencia {formato_monto(ciclo["cuota_referencia"])}.', 'success')
            else:
                flash(f'El ciclo {mes} ya estaba generado.', 'info')
        except Exception as exc:
            db.rollback()
            flash(f'No se pudo generar el ciclo de cobranza: {exc}', 'danger')
        return redirect_back('cuotas_view', mes=mes)

    @app.get('/pagos/<int:pago_id>/comprobante.pdf')
    @login_required
    def pagos_comprobante(pago_id: int):
        db = get_db()
        pago = db.fetchone(
            """
            SELECT p.id, p.fecha, p.mes, p.monto, p.observacion,
                   a.nombre AS parcela, a.curso AS sector, a.apoderado AS propietario,
                   a.direccion, a.telefono
            FROM pagos_parcelas p
            INNER JOIN parcelas a ON a.id = p.parcela_id
            WHERE p.id = ?
            """,
            (pago_id,),
        )
        if not pago:
            flash('Pago no encontrado.', 'danger')
            return redirect(url_for('pagos_list'))
        pdf = construir_pdf_comprobante_pago(pago)
        return send_file(pdf, mimetype='application/pdf', as_attachment=True, download_name=f'comprobante_pago_{pago_id}_{pago["mes"]}.pdf')

    @app.get('/reportes/mensual')
    @login_required
    def reportes_mensual():
        db = get_db()
        mes = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        resumen = construir_reporte_mensual(db, mes)
        return render_template('reportes_mensual.html', mes=mes, reporte=resumen)

    @app.get('/reportes/mensual/export/<fmt>')
    @login_required
    def reportes_mensual_export(fmt: str):
        db = get_db()
        mes = request.args.get('mes') or datetime.today().strftime('%Y-%m')
        reporte = construir_reporte_mensual(db, mes)
        if fmt == 'pdf':
            data = exportar_reporte_mensual_pdf(reporte)
            return send_file(data, mimetype='application/pdf', as_attachment=True, download_name=f'reporte_mensual_{mes}.pdf')
        if fmt == 'xlsx':
            data = exportar_reporte_mensual_xlsx(reporte)
            return send_file(data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'reporte_mensual_{mes}.xlsx')
        if fmt == 'csv':
            sio = StringIO()
            w = csv.writer(sio)
            w.writerow(['Mes', reporte['mes']])
            w.writerow(['Ingresos', reporte['ingresos']])
            w.writerow(['Gastos', reporte['gastos']])
            w.writerow(['Balance', reporte['balance']])
            w.writerow(['Esperado', reporte['total_esperado']])
            w.writerow(['Pagado', reporte['total_pagado']])
            w.writerow(['Pendiente', reporte['total_pendiente']])
            w.writerow([])
            w.writerow(['Parcela', 'Cuota', 'Pagado', 'Pendiente', 'Estado'])
            for fila in reporte['cuotas']:
                w.writerow([fila['nombre'], fila['cuota_mensual'], fila['pagado'], fila['pendiente'], fila['estado']])
            data = BytesIO(sio.getvalue().encode('utf-8-sig'))
            return send_file(data, mimetype='text/csv', as_attachment=True, download_name=f'reporte_mensual_{mes}.csv')
        flash('Formato no soportado.', 'danger')
        return redirect(url_for('reportes_mensual', mes=mes))

    return app


def is_postgres_url(url: str) -> bool:
    return url.startswith('postgresql://') or url.startswith('postgres://')


def sql_like_ci(value: str) -> str:
    return f"%{(value or '').strip().lower()}%"


def get_default_condominio_id(db: DBAdapter) -> int:
    row = db.fetchone('SELECT id FROM condominios ORDER BY id LIMIT 1')
    return int(row['id']) if row else 1


def get_current_condominio_id(db: DBAdapter) -> int | None:
    try:
        if getattr(current_user, 'is_authenticated', False):
            selected_condominio = request.values.get('condominio_id', '').strip()
            if getattr(current_user, 'is_global_admin', lambda: False)() and selected_condominio.isdigit():
                return int(selected_condominio)
            if getattr(current_user, 'condominio_id', None):
                return int(current_user.condominio_id)
            if getattr(current_user, 'role', None) == 'admin' and selected_condominio.isdigit():
                return int(selected_condominio)
    except Exception:
        pass
    return get_default_condominio_id(db)


def get_current_condominio(db: DBAdapter):
    condominio_id = get_current_condominio_id(db)
    if condominio_id is None:
        return None
    return db.fetchone('SELECT * FROM condominios WHERE id = ?', (condominio_id,))


def require_same_condominio(db: DBAdapter, table: str, row_id: int) -> bool:
    if getattr(current_user, 'is_authenticated', False) and getattr(current_user, 'is_global_admin', lambda: False)():
        return True
    condominio_id = get_current_condominio_id(db)
    row = db.fetchone(f'SELECT id FROM {table} WHERE id = ? AND condominio_id = ?', (row_id, condominio_id))
    return row is not None


def condominio_scope_clause(alias: str, db: DBAdapter) -> tuple[str, list[Any]]:
    if getattr(current_user, 'is_authenticated', False) and getattr(current_user, 'is_global_admin', lambda: False)():
        condominio_id = request.args.get('condominio_id', '').strip()
        if condominio_id.isdigit():
            return f' AND {alias}.condominio_id = ?', [int(condominio_id)]
        return '', []
    return f' AND {alias}.condominio_id = ?', [get_current_condominio_id(db)]


def obtener_movimientos_filtrados(db: DBAdapter, tipo: str = 'Todos', mes: str = '', q: str = '', fecha_desde: str = '', fecha_hasta: str = '', actividad_id: str | int = '', parcela_id: str | int = ''):
    sql = """
        SELECT m.id, m.fecha, m.tipo, m.concepto, m.monto, m.condominio_id, COALESCE(a.nombre, '-') AS actividad,
               COALESCE(al.nombre, '-') AS parcela,
               COALESCE(m.origen, 'general') AS origen, COALESCE(m.observacion, '') AS observacion
        FROM movimientos m
        LEFT JOIN actividades a ON a.id = m.actividad_id
        LEFT JOIN parcelas al ON al.id = m.parcela_id
        WHERE 1=1
    """
    params: list[Any] = []
    cond_clause, cond_params = condominio_scope_clause('m', db)
    sql += cond_clause
    params.extend(cond_params)
    if tipo in ('ingreso', 'gasto'):
        sql += ' AND m.tipo = ?'
        params.append(tipo)
    if mes:
        sql += ' AND substr(m.fecha, 1, 7) = ?'
        params.append(mes)
    if fecha_desde:
        sql += ' AND m.fecha >= ?'
        params.append(fecha_desde)
    if fecha_hasta:
        sql += ' AND m.fecha <= ?'
        params.append(fecha_hasta)
    if actividad_id:
        sql += ' AND m.actividad_id = ?'
        params.append(int(actividad_id))
    if parcela_id:
        sql += ' AND m.parcela_id = ?'
        params.append(int(parcela_id))
    if q:
        like = sql_like_ci(q)
        sql += " AND (LOWER(COALESCE(m.concepto, '')) LIKE ? OR LOWER(COALESCE(m.observacion, '')) LIKE ? OR LOWER(COALESCE(m.fecha, '')) LIKE ? OR LOWER(COALESCE(m.origen, '')) LIKE ? OR LOWER(COALESCE(a.nombre, '')) LIKE ? OR LOWER(COALESCE(al.nombre, '')) LIKE ? OR LOWER(COALESCE(al.curso, '')) LIKE ?)"
        params.extend([like, like, like, like, like, like, like])
    sql += ' ORDER BY m.fecha DESC, m.id DESC'
    return db.fetchall(sql, params)


def extraer_acta_form(request, db):
    condominio = get_current_condominio(db)
    nombre_condominio = condominio['nombre'] if condominio else 'Condominio'
    return {
        'titulo': request.form.get('titulo', '').strip() or f"Acta Asamblea {nombre_condominio}",
        'fecha': request.form.get('fecha', '').strip() or datetime.today().strftime('%Y-%m-%d'),
        'lugar': request.form.get('lugar', '').strip(),
        'hora_inicio': request.form.get('hora_inicio', '').strip(),
        'hora_termino': request.form.get('hora_termino', '').strip(),
        'asistentes': request.form.get('asistentes', '').strip(),
        'temas': request.form.get('temas', '').strip(),
        'desarrollo': request.form.get('desarrollo', '').strip(),
        'acuerdos': request.form.get('acuerdos', '').strip(),
        'responsables': request.form.get('responsables', '').strip(),
        'observaciones': request.form.get('observaciones', '').strip(),
        'estado': request.form.get('estado', 'borrador').strip() or 'borrador',
    }


def exportar_acta_pdf(acta) -> BytesIO:
    data = BytesIO()
    doc = SimpleDocTemplate(data, pagesize=A4, rightMargin=18*mm, leftMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph(f"<b>{acta['titulo']}</b>", styles['Title']))
    story.append(Spacer(1, 6))
    meta = [
        ['Fecha', acta['fecha'] or '-'],
        ['Lugar', acta['lugar'] or '-'],
        ['Hora de inicio', acta['hora_inicio'] or '-'],
        ['Hora de término', acta['hora_termino'] or '-'],
        ['Estado', str(acta['estado']).replace('_', ' ').title()],
    ]
    tbl = Table(meta, colWidths=[35*mm, 130*mm])
    tbl.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.whitesmoke),('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#d9d2c2')),('VALIGN',(0,0),(-1,-1),'TOP')]))
    story.append(tbl)
    story.append(Spacer(1, 10))
    for titulo, key in [('Asistentes', 'asistentes'), ('Temas', 'temas'), ('Desarrollo', 'desarrollo'), ('Acuerdos', 'acuerdos'), ('Responsables y plazos', 'responsables'), ('Observaciones', 'observaciones')]:
        story.append(Paragraph(f"<b>{titulo}</b>", styles['Heading3']))
        contenido = (acta[key] or '-').replace('\n', '<br/>')
        story.append(Paragraph(contenido, styles['BodyText']))
        story.append(Spacer(1, 8))
    doc.build(story)
    data.seek(0)
    return data


def seed_default_acta_modelo(db: DBAdapter) -> None:
    row = db.fetchone('SELECT COUNT(*) AS total FROM actas')
    total = int(row['total'] or 0) if row else 0
    if total:
        return
    db.execute(
        """
        INSERT INTO actas (titulo, fecha, lugar, hora_inicio, hora_termino, asistentes, temas, desarrollo, acuerdos, responsables, observaciones, estado, created_by, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            f"Acta Asamblea {get_current_condominio(db)['nombre'] if get_current_condominio(db) else 'Condominio'}",
            '2026-03-14',
            'Parcela T, Claudia Saavedra',
            '12:15',
            '',
            'Parcela A: Ramón Figueroa\nParcela B: Luis Ardizzi\nParcela C: Raúl Jopia\nParcela D: Luis Caro\nParcela G: Madeline Cruces\nParcela L: Francisco Astudillo\nParcela M: Alexis Pino\nParcela Q: Romina Pérez\nParcela R: Jesús Lizana\nParcela S: Igor Ledermann\nParcela T: Claudia Saavedra',
            '1. Cambio de cargos en directiva\n2. Revisión de cuentas\n3. Estatutos de Consejo Vecinal de Desarrollo (borrador)\n4. Evaluación tareas del jardinero\n5. Cambio de motor del portón de acceso\n6. Otros',
            'Se informa cambio de presidencia y se revisan ingresos, egresos, morosidad, estatutos, tareas del jardinero y necesidad de cambiar el motor del portón.',
            'Enviar saldo disponible.\nEnviar borrador de estatutos.\nRevisión por residentes.\nReunión con jardinero.\nCambio de motor.\nDesactivar aperturas a morosos con 2 o más meses.',
            'Tesorero: 17/03/2026\nDirectiva: 17/03/2026 y 31/03/2026\nTodos los residentes: 31/03/2026',
            'Nota: En caso de no remitir observaciones luego de 48 horas hábiles de su recepción, se dará por aprobada el acta.',
            'aprobada',
            'Sistema',
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        ),
    )
    db.commit()


def exportar_movimientos_pdf(movimientos, school_name: str, school_location: str, filtros: dict[str, str]) -> BytesIO:
    data = BytesIO()
    doc = SimpleDocTemplate(data, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    elems = []
    elems.append(Paragraph(f'{school_name} · {school_location}', styles['Title']))
    elems.append(Paragraph(f'Reporte de movimientos · generado {datetime.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
    filtros_txt = ' · '.join([f'{k}: {v}' for k, v in filtros.items()])
    elems.append(Paragraph(filtros_txt, styles['Normal']))
    elems.append(Spacer(1, 6))
    total_ing = sum(float(r['monto']) for r in movimientos if r['tipo'] == 'ingreso')
    total_gas = sum(float(r['monto']) for r in movimientos if r['tipo'] == 'gasto')
    elems.append(Paragraph(f'Registros: {len(movimientos)} · Ingresos: {formato_monto(total_ing)} · Gastos: {formato_monto(total_gas)} · Balance: {formato_monto(total_ing-total_gas)}', styles['Heading3']))
    table_data = [['Fecha', 'Tipo', 'Concepto', 'Categoría', 'Parcela', 'Origen', 'Monto']]
    for row in movimientos:
        concepto = str(row['concepto'])
        if len(concepto) > 38:
            concepto = concepto[:35] + '...'
        table_data.append([row['fecha'], row['tipo'], concepto, row['actividad'], (row['parcela'] if ('parcela' in row.keys()) else '-'), row['origen'], formato_monto(row['monto'])])
    table = Table(table_data, repeatRows=1, colWidths=[24*mm, 20*mm, 78*mm, 42*mm, 42*mm, 30*mm, 24*mm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#dfe7ff')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor('#cbd5e1')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (-1,1), (-1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8fafc')]),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
    ]))
    elems.append(table)
    doc.build(elems)
    data.seek(0)
    return data


def crear_backup_db(database_path: str | Path) -> Path:
    database_path = str(database_path)
    BACKUP_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if is_postgres_url(database_path):
        destino = BACKUP_DIR / f'backup_contabilidad_{timestamp}.sql'
        parsed = urlparse(database_path)
        env = os.environ.copy()
        if parsed.password:
            env['PGPASSWORD'] = parsed.password
        cmd = [
            'pg_dump',
            '-h', parsed.hostname or 'localhost',
            '-p', str(parsed.port or 5432),
            '-U', parsed.username or 'postgres',
            '-d', (parsed.path or '/').lstrip('/'),
            '-f', str(destino),
        ]
        try:
            subprocess.run(cmd, check=True, env=env, capture_output=True)
        except FileNotFoundError as exc:
            raise RuntimeError('pg_dump no está instalado o no está en el PATH.') from exc
        except subprocess.CalledProcessError as exc:
            raise RuntimeError(exc.stderr.decode('utf-8', errors='ignore') or 'pg_dump falló.') from exc
        return destino
    origen = Path(database_path)
    destino = BACKUP_DIR / f'backup_contabilidad_{timestamp}.db'
    shutil.copy2(origen, destino)
    return destino


def listar_backups():
    if not BACKUP_DIR.exists():
        return []
    archivos = sorted([*BACKUP_DIR.glob('*.db'), *BACKUP_DIR.glob('*.sql')], key=lambda p: p.stat().st_mtime, reverse=True)
    return [
        {
            'nombre': p.name,
            'tamano': p.stat().st_size,
            'modificado': datetime.fromtimestamp(p.stat().st_mtime),
        }
        for p in archivos
    ]


def formato_monto(valor: Any) -> str:
    try:
        valor = float(valor or 0)
    except Exception:
        valor = 0
    return f"${valor:,.0f}".replace(',', '.')


def parse_float(raw: str) -> float:
    txt = (raw or '').strip()
    if not txt:
        return 0.0
    txt = txt.replace('.', '').replace(',', '.') if ',' in txt else txt.replace(',', '.')
    return float(txt)


def validar_fecha(fecha: str) -> bool:
    datetime.strptime(fecha, '%Y-%m-%d')
    return True


def estado_cuota(cuota: Any, pagado: Any) -> tuple[str, str]:
    cuota_f = float(cuota or 0)
    pagado_f = float(pagado or 0)
    if cuota_f <= 0:
        return ('Sin cuota', '⚪')
    if pagado_f >= cuota_f:
        return ('Pagado', '🟢')
    if pagado_f > 0:
        return ('Parcial', '🟡')
    return ('Deuda', '🔴')


def obtener_nombre_parcela(db: DBAdapter, parcela_id: int) -> str:
    row = db.fetchone('SELECT nombre FROM parcelas WHERE id=? AND condominio_id = ?', (parcela_id, get_current_condominio_id(db)))
    return row['nombre'] if row else 'Parcela'


def parcela_duplicado(db: DBAdapter, nombre: str, curso: str, exclude_id: int | None = None) -> bool:
    sql = "SELECT id FROM parcelas WHERE condominio_id = ? AND lower(trim(nombre)) = lower(trim(?)) AND lower(trim(COALESCE(curso, ''))) = lower(trim(?))"
    params: list[Any] = [get_current_condominio_id(db), nombre, curso or '']
    if exclude_id:
        sql += ' AND id <> ?'
        params.append(exclude_id)
    return db.fetchone(sql, params) is not None


def pago_duplicado(db: DBAdapter, parcela_id: int, mes: str) -> bool:
    row = db.fetchone('SELECT id FROM pagos_parcelas WHERE parcela_id = ? AND mes = ? AND condominio_id = ?', (parcela_id, mes, get_current_condominio_id(db)))
    return row is not None


def registrar_pago_parcela(db: DBAdapter, parcela_id: int, fecha: str, mes: str, monto: float, observacion: str, actividad_id: int | None = None, tipo_pago: str = 'cuota_mensual') -> None:
    parcela = db.fetchone('SELECT * FROM parcelas WHERE id = ? AND condominio_id = ?', (parcela_id, get_current_condominio_id(db)))
    if not parcela:
        raise ValueError('Parcela no encontrada')
    if tipo_pago == 'cuota_mensual':
        concepto = f'Gasto común parcela: {parcela["nombre"]} ({mes})'
        if db.kind == 'postgres':
            cur = db.execute(
                'INSERT INTO movimientos (fecha, tipo, concepto, monto, actividad_id, parcela_id, observacion, origen, condominio_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?) RETURNING id',
                (fecha, 'ingreso', concepto, monto, None, parcela_id, observacion, 'cuota_mensual', get_current_condominio_id(db)),
            )
            movimiento_id = cur.fetchone()['id']
        else:
            cur = db.execute(
                'INSERT INTO movimientos (fecha, tipo, concepto, monto, actividad_id, parcela_id, observacion, origen, condominio_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
                (fecha, 'ingreso', concepto, monto, None, parcela_id, observacion, 'cuota_mensual', get_current_condominio_id(db)),
            )
            movimiento_id = cur.lastrowid
        db.execute(
            'INSERT INTO pagos_parcelas (parcela_id, fecha, mes, monto, observacion, movimiento_id, condominio_id) VALUES (?, ?, ?, ?, ?, ?, ?)',
            (parcela_id, fecha, mes, monto, observacion, movimiento_id, get_current_condominio_id(db)),
        )
    else:
        concepto = f'Ingreso extraordinario parcela: {parcela["nombre"]}'
        detalle = observacion if observacion else f'Ingreso extraordinario registrado en {mes}'
        db.execute(
            'INSERT INTO movimientos (fecha, tipo, concepto, monto, actividad_id, parcela_id, observacion, origen, condominio_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (fecha, 'ingreso', concepto, monto, actividad_id, parcela_id, detalle, 'actividad_parcela', get_current_condominio_id(db)),
        )


def resumen_cuotas_por_parcela(db: DBAdapter, mes: str):
    return db.fetchall(
        """
        SELECT a.id, a.nombre, a.curso, a.cuota_mensual, a.activo,
               COALESCE(SUM(CASE WHEN p.mes = ? THEN p.monto ELSE 0 END), 0) AS pagado
        FROM parcelas a
        LEFT JOIN pagos_parcelas p ON a.id = p.parcela_id AND p.condominio_id = a.condominio_id
        WHERE a.condominio_id = ?
        GROUP BY a.id, a.nombre, a.curso, a.cuota_mensual, a.activo
        ORDER BY a.nombre
        """,
        (mes, get_current_condominio_id(db)),
    )


def obtener_alertas_morosidad(db: DBAdapter, mes: str):
    alertas = []
    for fila in resumen_cuotas_por_parcela(db, mes):
        if not fila['activo']:
            continue
        debe = max(float(fila['cuota_mensual']) - float(fila['pagado']), 0)
        if debe > 0:
            estado, icono = estado_cuota(fila['cuota_mensual'], fila['pagado'])
            alertas.append({
                'parcela_id': fila['id'],
                'nombre': fila['nombre'],
                'curso': fila['curso'],
                'debe': debe,
                'estado': estado,
                'icono': icono,
            })
    return alertas


def meses_hasta_corte(mes_corte: str) -> list[str]:
    corte = datetime.strptime(mes_corte + '-01', '%Y-%m-%d')
    mes_inicio = 3
    if corte.month < mes_inicio:
        return []
    return [f"{corte.year}-{mes:02d}" for mes in range(mes_inicio, corte.month + 1)]


def nombre_mes_es(numero_mes: int) -> str:
    nombres = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril', 5: 'mayo', 6: 'junio',
        7: 'julio', 8: 'agosto', 9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre',
    }
    return nombres.get(numero_mes, str(numero_mes))


def resumen_deuda_acumulada_por_parcela(db: DBAdapter, mes_corte: str):
    meses = meses_hasta_corte(mes_corte)
    filas = db.fetchall(
        """
        SELECT a.id, a.nombre, a.curso, a.cuota_mensual, a.activo
        FROM parcelas a
        ORDER BY a.nombre
        """
    )
    pagos_rows = db.fetchall(
        """
        SELECT p.parcela_id, p.mes, COALESCE(SUM(p.monto), 0) AS monto
        FROM pagos_parcelas p
        WHERE substr(p.mes, 1, 4) = ? AND p.mes <= ?
        GROUP BY p.parcela_id, p.mes
        """,
        (mes_corte[:4], mes_corte),
    )
    pagos_map: dict[tuple[int, str], float] = {}
    for row in pagos_rows:
        pagos_map[(int(row['parcela_id']), row['mes'])] = float(row['monto'] or 0)

    resumen = []
    for fila in filas:
        cuota = float(fila['cuota_mensual'] or 0)
        detalle_deuda = []
        pagado_acumulado = 0.0
        deuda_total = 0.0
        for mes in meses:
            pagado_mes = float(pagos_map.get((int(fila['id']), mes), 0) or 0)
            pagado_acumulado += pagado_mes
            deuda_mes = max(cuota - pagado_mes, 0) if fila['activo'] else 0.0
            deuda_total += deuda_mes
            if fila['activo'] and deuda_mes > 0:
                detalle_deuda.append({
                    'mes': mes,
                    'deuda': deuda_mes,
                })

        esperado_acumulado = cuota * len(meses) if fila['activo'] else 0.0
        resumen.append({
            'id': fila['id'],
            'nombre': fila['nombre'],
            'curso': fila['curso'],
            'cuota_mensual': cuota,
            'activo': fila['activo'],
            'meses_considerados': len(meses),
            'esperado_acumulado': esperado_acumulado,
            'pagado_acumulado': pagado_acumulado,
            'deuda_total': deuda_total,
            'detalle_deuda': detalle_deuda,
        })
    return resumen


def construir_pdf_deudores(db: DBAdapter, mes_corte: str, modo: str = 'deuda') -> BytesIO:
    filas = resumen_deuda_acumulada_por_parcela(db, mes_corte)
    if modo == 'deuda':
        filas = [fila for fila in filas if fila['activo'] and fila['deuda_total'] > 0]
    else:
        filas = [fila for fila in filas if fila['activo']]

    total_esperado = sum(float(f['esperado_acumulado']) for f in filas)
    total_pagado = sum(float(f['pagado_acumulado']) for f in filas)
    total_deuda = sum(float(f['deuda_total']) for f in filas)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    elements = []
    corte_dt = datetime.strptime(mes_corte + '-01', '%Y-%m-%d')
    titulo = 'Reporte de parcelas con deuda' if modo == 'deuda' else 'Reporte general de parcelas'
    filtro_txt = 'solo con deuda' if modo == 'deuda' else 'todos'
    subtitulo = f'Deuda acumulada desde enero hasta {nombre_mes_es(corte_dt.month)} de {corte_dt.year} · filtro: {filtro_txt}'

    elements.append(Paragraph(f'<b>{SCHOOL_NAME}</b>', styles['Title']))
    elements.append(Paragraph(titulo, styles['Heading2']))
    elements.append(Paragraph(subtitulo, styles['Normal']))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(
        f'Parcelas incluidas: {len(filas)} &nbsp;&nbsp;&nbsp; Total esperado: {formato_monto(total_esperado)} &nbsp;&nbsp;&nbsp; Total pagado: {formato_monto(total_pagado)} &nbsp;&nbsp;&nbsp; Deuda total: {formato_monto(total_deuda)}',
        styles['Normal']
    ))
    elements.append(Spacer(1, 8))

    data = [['Parcela', 'Sector', 'Gasto común', 'Meses', 'Esperado', 'Pagado', 'Debe', 'Meses adeudados']]
    for fila in filas:
        meses_adeudados = ', '.join(nombre_mes_es(int(item['mes'][5:7])) for item in fila['detalle_deuda']) or 'Sin deuda'
        data.append([
            fila['nombre'],
            fila['curso'] or '-',
            formato_monto(fila['cuota_mensual']),
            str(fila['meses_considerados']),
            formato_monto(fila['esperado_acumulado']),
            formato_monto(fila['pagado_acumulado']),
            formato_monto(fila['deuda_total']),
            meses_adeudados,
        ])

    table = Table(data, repeatRows=1, colWidths=[60 * mm, 28 * mm, 23 * mm, 15 * mm, 26 * mm, 26 * mm, 24 * mm, 68 * mm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor('#f8fafc')]),
        ('ALIGN', (2, 1), (6, -1), 'RIGHT'),
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer



def obtener_ciclo_cobranza(db: DBAdapter, mes: str):
    row = db.fetchone('SELECT mes, cuota_referencia, generado_en, generado_por FROM ciclos_cobranza WHERE mes = ? AND condominio_id = ?', (mes, get_current_condominio_id(db)))
    if not row:
        return None
    return {
        'mes': row['mes'],
        'cuota_referencia': float(row['cuota_referencia'] or 0),
        'generado_en': row['generado_en'],
        'generado_por': row['generado_por'],
    }


def generar_ciclo_cobranza(db: DBAdapter, mes: str, generado_por: str):
    existente = obtener_ciclo_cobranza(db, mes)
    cuota_ref_row = db.fetchone('SELECT COALESCE(AVG(cuota_mensual), 40000) AS cuota FROM parcelas WHERE activo = 1 AND condominio_id = ?', (get_current_condominio_id(db),))
    cuota_ref = float(cuota_ref_row['cuota'] or 40000)
    if existente:
        return {'creado': False, 'cuota_referencia': cuota_ref}
    generado_en = datetime.now().strftime('%Y-%m-%d %H:%M')
    db.execute(
        'INSERT INTO ciclos_cobranza (mes, cuota_referencia, generado_en, generado_por, observacion, condominio_id) VALUES (?, ?, ?, ?, ?, ?)',
        (mes, cuota_ref, generado_en, generado_por, 'Ciclo mensual generado desde el panel.', get_current_condominio_id(db)),
    )
    return {'creado': True, 'cuota_referencia': cuota_ref}


def construir_pdf_comprobante_pago(pago) -> BytesIO:
    data = BytesIO()
    doc = SimpleDocTemplate(data, pagesize=A4, leftMargin=16*mm, rightMargin=16*mm, topMargin=18*mm, bottomMargin=18*mm)
    styles = getSampleStyleSheet()
    elems = []
    elems.append(Paragraph(f'<b>{SCHOOL_NAME}</b>', styles['Title']))
    elems.append(Paragraph('Comprobante de pago de gasto común', styles['Heading2']))
    elems.append(Paragraph(f'Emitido: {datetime.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
    elems.append(Spacer(1, 8))
    tabla = [
        ['Parcela', pago['parcela']],
        ['Sector', pago['sector'] or '-'],
        ['Propietario / contacto', pago['propietario'] or '-'],
        ['Fecha de pago', pago['fecha']],
        ['Mes abonado', pago['mes']],
        ['Monto', formato_monto(pago['monto'])],
        ['Observación', pago['observacion'] or '-'],
    ]
    t = Table(tabla, colWidths=[52*mm, 118*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#eef4e7')),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.35, colors.HexColor('#cfd8c2')),
        ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#aab89a')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    elems.append(t)
    elems.append(Spacer(1, 10))
    elems.append(Paragraph('Este comprobante acredita el ingreso registrado en el sistema administrativo del condominio.', styles['Italic']))
    doc.build(elems)
    data.seek(0)
    return data


def construir_reporte_mensual(db: DBAdapter, mes: str) -> dict[str, Any]:
    condominio_id = get_current_condominio_id(db)
    resumen_mov = db.fetchone(
        """
        SELECT COALESCE(SUM(CASE WHEN tipo='ingreso' THEN monto ELSE 0 END),0) AS ingresos,
               COALESCE(SUM(CASE WHEN tipo='gasto' THEN monto ELSE 0 END),0) AS gastos,
               COUNT(*) AS movimientos
        FROM movimientos
        WHERE substr(fecha,1,7) = ?
          AND condominio_id = ?
        """,
        (mes, condominio_id),
    )
    cuotas = []
    for fila in resumen_cuotas_por_parcela(db, mes):
        if not fila['activo']:
            continue
        pendiente = max(float(fila['cuota_mensual']) - float(fila['pagado']), 0)
        estado, _ = estado_cuota(fila['cuota_mensual'], fila['pagado'])
        cuotas.append({
            'id': fila['id'],
            'nombre': fila['nombre'],
            'curso': fila['curso'],
            'cuota_mensual': float(fila['cuota_mensual'] or 0),
            'pagado': float(fila['pagado'] or 0),
            'pendiente': pendiente,
            'estado': estado,
        })
    cuotas.sort(key=lambda x: (x['pendiente'] == 0, x['nombre']))
    morosos = [x for x in cuotas if x['pendiente'] > 0]
    top_gastos = db.fetchall(
        """
        SELECT COALESCE(a.nombre, 'Sin categoría') AS categoria, COALESCE(SUM(m.monto),0) AS total
        FROM movimientos m
        LEFT JOIN actividades a ON a.id = m.actividad_id
        WHERE m.tipo = 'gasto' AND substr(m.fecha,1,7) = ? AND m.condominio_id = ?
        GROUP BY COALESCE(a.nombre, 'Sin categoría')
        ORDER BY total DESC, categoria ASC
        LIMIT 8
        """,
        (mes, condominio_id),
    )
    pagos = db.fetchall(
        """
        SELECT p.id, p.fecha, p.mes, p.monto, a.nombre AS parcela
        FROM pagos_parcelas p
        INNER JOIN parcelas a ON a.id = p.parcela_id
        WHERE p.mes = ? AND p.condominio_id = ?
        ORDER BY p.fecha DESC, a.nombre
        """,
        (mes, condominio_id),
    )
    total_esperado = sum(x['cuota_mensual'] for x in cuotas)
    total_pagado = sum(x['pagado'] for x in cuotas)
    total_pendiente = sum(x['pendiente'] for x in cuotas)
    ciclo = obtener_ciclo_cobranza(db, mes)
    return {
        'mes': mes,
        'ingresos': float(resumen_mov['ingresos'] or 0),
        'gastos': float(resumen_mov['gastos'] or 0),
        'balance': float(resumen_mov['ingresos'] or 0) - float(resumen_mov['gastos'] or 0),
        'movimientos': int(resumen_mov['movimientos'] or 0),
        'total_esperado': total_esperado,
        'total_pagado': total_pagado,
        'total_pendiente': total_pendiente,
        'cumplimiento': round((total_pagado / total_esperado) * 100, 1) if total_esperado else 100.0,
        'cuotas': cuotas,
        'morosos': morosos,
        'top_gastos': [dict(x) for x in top_gastos],
        'pagos': [dict(x) for x in pagos],
        'ciclo': ciclo,
    }


def exportar_reporte_mensual_pdf(reporte: dict[str, Any]) -> BytesIO:
    data = BytesIO()
    doc = SimpleDocTemplate(data, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    elems = []
    elems.append(Paragraph(f'{SCHOOL_NAME}', styles['Title']))
    elems.append(Paragraph(f'Reporte mensual · {reporte["mes"]}', styles['Heading2']))
    elems.append(Paragraph(f'Ingresos: {formato_monto(reporte["ingresos"])} · Gastos: {formato_monto(reporte["gastos"])} · Balance: {formato_monto(reporte["balance"])} · Cumplimiento: {reporte["cumplimiento"]}%', styles['Normal']))
    if reporte['ciclo']:
        elems.append(Paragraph(f'Ciclo generado por {reporte["ciclo"]["generado_por"]} el {reporte["ciclo"]["generado_en"]}.', styles['Normal']))
    elems.append(Spacer(1, 8))
    data_table = [['Parcela', 'Gasto común', 'Pagado', 'Pendiente', 'Estado']]
    for fila in reporte['cuotas']:
        data_table.append([fila['nombre'], formato_monto(fila['cuota_mensual']), formato_monto(fila['pagado']), formato_monto(fila['pendiente']), fila['estado']])
    table = Table(data_table, repeatRows=1, colWidths=[70*mm, 28*mm, 28*mm, 28*mm, 26*mm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#34421f')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.3, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.HexColor('#f8fafc')]),
        ('ALIGN', (1,1), (3,-1), 'RIGHT'),
    ]))
    elems.append(table)
    doc.build(elems)
    data.seek(0)
    return data


def exportar_reporte_mensual_xlsx(reporte: dict[str, Any]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte mensual'
    ws.append(['Mes', reporte['mes']])
    ws.append(['Ingresos', reporte['ingresos']])
    ws.append(['Gastos', reporte['gastos']])
    ws.append(['Balance', reporte['balance']])
    ws.append(['Esperado', reporte['total_esperado']])
    ws.append(['Pagado', reporte['total_pagado']])
    ws.append(['Pendiente', reporte['total_pendiente']])
    ws.append(['Cumplimiento %', reporte['cumplimiento']])
    ws.append([])
    ws.append(['Parcela', 'Gasto común', 'Pagado', 'Pendiente', 'Estado'])
    for fila in reporte['cuotas']:
        ws.append([fila['nombre'], fila['cuota_mensual'], fila['pagado'], fila['pendiente'], fila['estado']])
    for c in ws[10]:
        c.font = c.font.copy(bold=True)
    for col in ['A','B','C','D','E']:
        ws.column_dimensions[col].width = 22
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    return data



def seed_default_condominio(db: DBAdapter) -> None:
    db.execute("CREATE TABLE IF NOT EXISTS condominios (id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL UNIQUE, direccion TEXT, activo INTEGER NOT NULL DEFAULT 1)") if db.kind == 'sqlite' else db.execute("CREATE TABLE IF NOT EXISTS condominios (id BIGSERIAL PRIMARY KEY, nombre TEXT NOT NULL UNIQUE, direccion TEXT, activo INTEGER NOT NULL DEFAULT 1)")
    row = db.fetchone('SELECT id FROM condominios ORDER BY id LIMIT 1')
    if row:
        return
    db.execute('INSERT INTO condominios (nombre, direccion, activo) VALUES (?, ?, ?)', (DEFAULT_CONDOMINIO_NAME, '', 1))
    db.commit()

def seed_default_admin(db: DBAdapter) -> None:
    has_user = db.fetchone('SELECT 1 FROM usuarios LIMIT 1')
    if has_user:
        return
    username = os.environ.get('ADMIN_USER', 'admin')
    password = os.environ.get('ADMIN_PASSWORD', 'admin123')
    nombre = os.environ.get('ADMIN_NAME', 'Administrador')
    db.execute(
        'INSERT INTO usuarios (username, password_hash, role, nombre, activo, condominio_id, must_change_password) VALUES (?, ?, ?, ?, 1, NULL, 0)',
        (username, generate_password_hash(password), 'admin', nombre),
    )
    db.commit()


def seed_default_parcelas(db: DBAdapter) -> None:
    total = db.fetchone('SELECT COUNT(*) AS total FROM parcelas')
    letras = list('ABCDEFGHIJKLMNOPQRSTU')
    if total and int(total['total'] or 0) > 0:
        return
    for letra in letras:
        db.execute(
            'INSERT INTO parcelas (nombre, curso, cuota_mensual, apoderado, telefono, direccion, observacion_ficha, activo, condominio_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (f'Parcela {letra}', '', 40000, '', '', '', 'Completa aquí los datos del propietario y observaciones de la parcela.', 1, get_default_condominio_id(db)),
        )
    db.commit()


def seed_default_actividades(db: DBAdapter) -> None:
    existentes = db.fetchone('SELECT COUNT(*) AS total FROM actividades')
    if existentes and int(existentes['total'] or 0) > 0:
        return
    categorias = [
        ('Agua', '2026-01-01', 'Consumo y distribución de agua'),
        ('Luz áreas comunes', '2026-01-01', 'Electricidad y alumbrado común'),
        ('Mantención general', '2026-01-01', 'Mantenciones preventivas y correctivas'),
        ('Portón y accesos', '2026-01-01', 'Automatización, controles y acceso'),
        ('Jardinería y áreas verdes', '2026-01-01', 'Manejo de jardines y olivos'),
        ('Seguridad', '2026-01-01', 'Guardias, cámaras o monitoreo'),
        ('Administración', '2026-01-01', 'Servicios administrativos del condominio'),
        ('Aseo y retiro', '2026-01-01', 'Limpieza y retiro de residuos'),
        ('Reparaciones', '2026-01-01', 'Reparaciones no programadas'),
        ('Multas e intereses', '2026-01-01', 'Cargos por atraso o multas'),
        ('Fondo de reserva', '2026-01-01', 'Ahorro para contingencias y proyectos'),
    ]
    for nombre, fecha, descripcion in categorias:
        db.execute('INSERT INTO actividades (nombre, fecha, descripcion, condominio_id) VALUES (?, ?, ?, ?)', (nombre, fecha, descripcion, get_default_condominio_id(db)))
    db.commit()



def cleanup_demo_environment(db: DBAdapter) -> None:
    demo_condominios = ['Parcelas Los Olivos', 'Parcelas Alto Maitenes', 'Parcelas Vista Campo (Demo)']
    demo_users = ['demo@parcelia.cl', 'comite@parcelia.cl']

    for username in demo_users:
        db.execute('DELETE FROM usuarios WHERE lower(username)=?', (username,))

    condo_rows = db.fetchall('SELECT id FROM condominios WHERE nombre IN (?, ?, ?)', tuple(demo_condominios))
    condo_ids = [int(row['id']) for row in condo_rows]
    for condominio_id in condo_ids:
        for table in ['pagos_parcelas', 'movimientos', 'actas', 'parcelas', 'actividades', 'ciclos_cobranza']:
            if column_exists(db, table, 'condominio_id'):
                db.execute(f'DELETE FROM {table} WHERE condominio_id=?', (condominio_id,))
        db.execute('DELETE FROM condominios WHERE id=?', (condominio_id,))
    db.commit()


def seed_demo_environment(db: DBAdapter) -> None:
    demo_user = db.fetchone('SELECT id FROM usuarios WHERE lower(username)=?', ('demo@parcelia.cl',))
    if demo_user:
        return

    demo_condominios = [
        ('Parcelas Vista Campo (Demo)', 'Camino El Molino km 12, Chicureo'),
    ]
    condominio_ids: dict[str, int] = {}
    for nombre, direccion in demo_condominios:
        row = db.fetchone('SELECT id FROM condominios WHERE lower(nombre)=lower(?)', (nombre,))
        if not row:
            db.execute('INSERT INTO condominios (nombre, direccion, activo) VALUES (?, ?, 1)', (nombre, direccion))
            db.commit()
            row = db.fetchone('SELECT id FROM condominios WHERE lower(nombre)=lower(?)', (nombre,))
        condominio_ids[nombre] = int(row['id'])

    demo_cuota = 85000
    parcelas_demo = [
        ('Parcela 01', 'Sector Robles', 'Familia González', '+56 9 7812 3401'),
        ('Parcela 02', 'Sector Robles', 'María Paz Soto', '+56 9 6123 4550'),
        ('Parcela 03', 'Sector Robles', 'Inversiones El Maitén', '+56 9 8221 9921'),
        ('Parcela 04', 'Sector Quebrada', 'Patricio Vera', '+56 9 7331 1180'),
        ('Parcela 05', 'Sector Quebrada', 'Daniela Pizarro', '+56 9 9011 7420'),
        ('Parcela 06', 'Sector Quebrada', 'Rodrigo León', '+56 9 9899 5501'),
        ('Parcela 07', 'Sector Laguna', 'Camila Espinoza', '+56 9 8455 9920'),
        ('Parcela 08', 'Sector Laguna', 'Familia Araya', '+56 9 7001 5530'),
        ('Parcela 09', 'Sector Laguna', 'Ana Luisa Morales', '+56 9 6211 4100'),
        ('Parcela 10', 'Sector Mirador', 'Jorge Contreras', '+56 9 6120 3200'),
        ('Parcela 11', 'Sector Mirador', 'Constructora Vértice', '+56 9 7881 2401'),
        ('Parcela 12', 'Sector Mirador', 'Paula Villarroel', '+56 9 9112 6570'),
    ]

    for nombre, sector, propietario, telefono in parcelas_demo:
        row = db.fetchone('SELECT id FROM parcelas WHERE lower(nombre)=lower(?) AND condominio_id=?', (nombre, condominio_ids['Parcelas Vista Campo (Demo)']))
        if not row:
            db.execute(
                'INSERT INTO parcelas (nombre, curso, cuota_mensual, apoderado, telefono, direccion, observacion_ficha, activo, condominio_id) VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)',
                (nombre, sector, demo_cuota, propietario, telefono, 'Parcelación con acceso controlado', f'Propietario principal: {propietario}. Demo de Parcelia.', condominio_ids['Parcelas Vista Campo (Demo)'])
            )
    db.commit()

    actividades = [
        ('Gasto común', 'Cobranza mensual ordinaria'),
        ('Agua de pozo', 'Consumo y mantención del sistema de agua'),
        ('Guardia y portería', 'Control de acceso y turnos de seguridad'),
        ('Mantención caminos', 'Nivelación y reparación de caminos interiores'),
        ('Áreas verdes', 'Riego y mantención de zonas comunes'),
        ('Portón eléctrico', 'Automatización y soporte de accesos'),
        ('Fondo de reserva', 'Reserva para imprevistos y proyectos'),
    ]
    actividad_ids: dict[str, int] = {}
    for nombre, descripcion in actividades:
        row = db.fetchone('SELECT id FROM actividades WHERE lower(nombre)=lower(?) AND condominio_id=?', (nombre, condominio_ids['Parcelas Vista Campo (Demo)']))
        if not row:
            db.execute('INSERT INTO actividades (nombre, fecha, descripcion, condominio_id) VALUES (?, ?, ?, ?)', (nombre, '2026-01-01', descripcion, condominio_ids['Parcelas Vista Campo (Demo)']))
            db.commit()
            row = db.fetchone('SELECT id FROM actividades WHERE lower(nombre)=lower(?) AND condominio_id=?', (nombre, condominio_ids['Parcelas Vista Campo (Demo)']))
        actividad_ids[nombre] = int(row['id'])

    parcelas_rows = db.fetchall('SELECT id, nombre FROM parcelas WHERE condominio_id=? ORDER BY id', (condominio_ids['Parcelas Vista Campo (Demo)'],))
    parcela_ids = {row['nombre']: int(row['id']) for row in parcelas_rows}

    months = [
        ('2025-11', '2025-11-05', '2025-11-18', 10),
        ('2025-12', '2025-12-05', '2025-12-18', 10),
        ('2026-01', '2026-01-05', '2026-01-19', 11),
        ('2026-02', '2026-02-05', '2026-02-20', 11),
        ('2026-03', '2026-03-05', '2026-03-19', 10),
        ('2026-04', '2026-04-05', '2026-04-16', 9),
    ]
    cuotas_por_mes = {
        '2025-11': 82000,
        '2025-12': 82000,
        '2026-01': 85000,
        '2026-02': 85000,
        '2026-03': 85000,
        '2026-04': 87000,
    }

    pago_idx = 0
    pagos_insertados = 0
    for mes, fecha_ingreso, fecha_reserva, paid_count in months:
        if not db.fetchone('SELECT 1 FROM ciclos_cobranza WHERE mes=? AND condominio_id=?', (mes, condominio_ids['Parcelas Vista Campo (Demo)'])):
            db.execute(
                'INSERT INTO ciclos_cobranza (mes, cuota_referencia, generado_en, generado_por, observacion, condominio_id) VALUES (?, ?, ?, ?, ?, ?)',
                (mes, cuotas_por_mes[mes], f'{mes}-01 09:00', 'Sistema demo Parcelia', 'Ciclo generado automáticamente para demostración comercial.', condominio_ids['Parcelas Vista Campo (Demo)'])
            )

        if not db.fetchone('SELECT 1 FROM movimientos WHERE condominio_id=? AND concepto=? AND fecha=?', (condominio_ids['Parcelas Vista Campo (Demo)'], f'Cobranza gasto común {mes}', fecha_ingreso)):
            db.execute(
                'INSERT INTO movimientos (fecha, tipo, concepto, monto, actividad_id, observacion, origen, condominio_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                (fecha_ingreso, 'ingreso', f'Cobranza gasto común {mes}', cuotas_por_mes[mes] * paid_count, actividad_ids['Gasto común'], f'{paid_count} parcelas pagadas o abonadas en el periodo.', 'cuota', condominio_ids['Parcelas Vista Campo (Demo)'])
            )
            pagos_insertados += 1

        if not db.fetchone('SELECT 1 FROM movimientos WHERE condominio_id=? AND concepto=? AND fecha=?', (condominio_ids['Parcelas Vista Campo (Demo)'], f'Aporte fondo de reserva {mes}', fecha_reserva)):
            db.execute(
                'INSERT INTO movimientos (fecha, tipo, concepto, monto, actividad_id, observacion, origen, condominio_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                (fecha_reserva, 'ingreso', f'Aporte fondo de reserva {mes}', 150000, actividad_ids['Fondo de reserva'], 'Aporte mensual destinado a contingencias y mejoras.', 'general', condominio_ids['Parcelas Vista Campo (Demo)'])
            )

    gastos = [
        ('2025-11-08', 'Mantención caminos interiores', 280000, 'Mantención caminos', 'Perfilado después de lluvias tempranas.'),
        ('2025-11-14', 'Servicio de guardia de fin de semana', 190000, 'Guardia y portería', 'Cobertura reforzada por mayor flujo de visitas.'),
        ('2025-12-10', 'Reparación de bomba de pozo', 420000, 'Agua de pozo', 'Cambio de sello mecánico y revisión eléctrica.'),
        ('2025-12-16', 'Mantención portón principal', 165000, 'Portón eléctrico', 'Ajuste de motor y fotoceldas.'),
        ('2026-01-09', 'Nivelación de camino sector Laguna', 310000, 'Mantención caminos', 'Mejora de tránsito interno.'),
        ('2026-01-22', 'Riego y poda de áreas verdes', 140000, 'Áreas verdes', 'Mantención de zonas comunes.'),
        ('2026-02-07', 'Servicio mensual de seguridad', 220000, 'Guardia y portería', 'Turnos de control de acceso.'),
        ('2026-02-13', 'Análisis de agua y cloración', 125000, 'Agua de pozo', 'Control preventivo del sistema.'),
        ('2026-03-12', 'Reposición de luminarias acceso', 98000, 'Portón eléctrico', 'Cambio de luminarias y cableado menor.'),
        ('2026-03-26', 'Limpieza de acequias internas', 132000, 'Áreas verdes', 'Preparación para temporada seca.'),
        ('2026-04-04', 'Reparación de portón peatonal', 118000, 'Portón eléctrico', 'Ajuste de cerradura y brazo hidráulico.'),
        ('2026-04-11', 'Mantención extraordinaria de caminos', 355000, 'Mantención caminos', 'Refuerzo previo a invierno.'),
    ]
    for fecha, concepto, monto, actividad, observacion in gastos:
        if not db.fetchone('SELECT 1 FROM movimientos WHERE condominio_id=? AND concepto=? AND fecha=?', (condominio_ids['Parcelas Vista Campo (Demo)'], concepto, fecha)):
            db.execute(
                'INSERT INTO movimientos (fecha, tipo, concepto, monto, actividad_id, observacion, origen, condominio_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?)',
                (fecha, 'gasto', concepto, monto, actividad_ids[actividad], observacion, 'general', condominio_ids['Parcelas Vista Campo (Demo)'])
            )

    # Registrar pagos individuales de abril para mostrar morosidad realista
    pago_plan = [87000, 87000, 87000, 87000, 87000, 87000, 87000, 43500, 43500, 0, 0, 0]
    fecha_pago = '2026-04-15'
    for idx, parcela in enumerate(parcelas_rows):
        monto = pago_plan[idx] if idx < len(pago_plan) else 0
        if monto <= 0:
            continue
        exists = db.fetchone('SELECT 1 FROM pagos_parcelas WHERE parcela_id=? AND mes=?', (int(parcela['id']), '2026-04'))
        if exists:
            continue
        movimiento = db.fetchone('SELECT id FROM movimientos WHERE condominio_id=? AND concepto=? AND fecha=?', (condominio_ids['Parcelas Vista Campo (Demo)'], 'Cobranza gasto común 2026-04', '2026-04-05'))
        movimiento_id = int(movimiento['id']) if movimiento else None
        db.execute(
            'INSERT INTO pagos_parcelas (parcela_id, fecha, mes, monto, observacion, movimiento_id, condominio_id) VALUES (?, ?, ?, ?, ?, ?, ?)',
            (int(parcela['id']), fecha_pago, '2026-04', monto, 'Pago registrado para demostración comercial.', movimiento_id, condominio_ids['Parcelas Vista Campo (Demo)'])
        )

    if not db.fetchone('SELECT 1 FROM actas WHERE titulo=? AND condominio_id=?', ('Reunión comité abril', condominio_ids['Parcelas Vista Campo (Demo)'])):
        db.execute(
            'INSERT INTO actas (titulo, fecha, lugar, hora_inicio, hora_termino, asistentes, temas, desarrollo, acuerdos, responsables, observaciones, estado, created_by, updated_at, condominio_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (
                'Reunión comité abril', '2026-04-10', 'Sala multiuso acceso principal', '19:30', '21:00',
                'Administrador, comité de administración y 6 propietarios',
                'Morosidad, mantención de caminos, presupuesto de invierno',
                'Se revisó la cobranza de abril y el plan de mantención preventiva previo al invierno.',
                '1) Reforzar seguimiento a 3 parcelas morosas. 2) Aprobar mantención extraordinaria de caminos. 3) Cotizar cámaras para segundo acceso.',
                'Administración y comité',
                'Acta cargada como parte de la demo.', 'aprobada', 'Sistema demo Parcelia', '2026-04-10 21:05', condominio_ids['Parcelas Vista Campo (Demo)']
            )
        )

    # Usuario demo y usuario comité
    if not db.fetchone('SELECT 1 FROM usuarios WHERE lower(username)=?', ('demo@parcelia.cl',)):
        db.execute(
            'INSERT INTO usuarios (username, password_hash, role, nombre, activo, condominio_id, must_change_password) VALUES (?, ?, ?, ?, 1, ?, 0)',
            ('demo@parcelia.cl', generate_password_hash('123456'), 'admin', 'Demo Parcelia', condominio_ids['Parcelas Vista Campo (Demo)'])
        )
    if not db.fetchone('SELECT 1 FROM usuarios WHERE lower(username)=?', ('comite@parcelia.cl',)):
        db.execute(
            'INSERT INTO usuarios (username, password_hash, role, nombre, activo, condominio_id, must_change_password) VALUES (?, ?, ?, ?, 1, ?, 0)',
            ('comite@parcelia.cl', generate_password_hash('123456'), 'comite', 'Comité Parcelia', condominio_ids['Parcelas Vista Campo (Demo)'])
        )

    db.commit()

def init_db(db: DBAdapter) -> None:
    if db.kind == 'postgres':
        script = """
        CREATE TABLE IF NOT EXISTS actividades (
            id BIGSERIAL PRIMARY KEY,
            nombre TEXT NOT NULL,
            fecha TEXT,
            descripcion TEXT
        );

        CREATE TABLE IF NOT EXISTS movimientos (
            id BIGSERIAL PRIMARY KEY,
            fecha TEXT NOT NULL,
            tipo TEXT NOT NULL CHECK(tipo IN ('ingreso', 'gasto')),
            concepto TEXT NOT NULL,
            monto DOUBLE PRECISION NOT NULL CHECK(monto >= 0),
            actividad_id BIGINT,
            parcela_id BIGINT,
            observacion TEXT,
            origen TEXT NOT NULL DEFAULT 'general',
            CONSTRAINT fk_mov_actividad FOREIGN KEY (actividad_id) REFERENCES actividades(id)
        );

        CREATE TABLE IF NOT EXISTS parcelas (
            id BIGSERIAL PRIMARY KEY,
            nombre TEXT NOT NULL,
            curso TEXT,
            cuota_mensual DOUBLE PRECISION NOT NULL DEFAULT 0,
            apoderado TEXT,
            telefono TEXT,
            direccion TEXT,
            observacion_ficha TEXT,
            activo INTEGER NOT NULL DEFAULT 1,
            must_change_password INTEGER NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS pagos_parcelas (
            id BIGSERIAL PRIMARY KEY,
            parcela_id BIGINT NOT NULL,
            fecha TEXT NOT NULL,
            mes TEXT NOT NULL,
            monto DOUBLE PRECISION NOT NULL CHECK(monto >= 0),
            observacion TEXT,
            movimiento_id BIGINT,
            CONSTRAINT fk_pagos_parcela FOREIGN KEY (parcela_id) REFERENCES parcelas(id),
            CONSTRAINT fk_pagos_mov FOREIGN KEY (movimiento_id) REFERENCES movimientos(id)
        );

        CREATE TABLE IF NOT EXISTS usuarios (
            id BIGSERIAL PRIMARY KEY,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            nombre TEXT NOT NULL,
            activo INTEGER NOT NULL DEFAULT 1,
            must_change_password INTEGER NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS ciclos_cobranza (
            mes TEXT PRIMARY KEY,
            cuota_referencia DOUBLE PRECISION NOT NULL DEFAULT 0,
            generado_en TEXT NOT NULL,
            generado_por TEXT NOT NULL,
            observacion TEXT
        );

        CREATE TABLE IF NOT EXISTS actas (
            id BIGSERIAL PRIMARY KEY,
            titulo TEXT NOT NULL,
            fecha TEXT NOT NULL,
            lugar TEXT,
            hora_inicio TEXT,
            hora_termino TEXT,
            asistentes TEXT,
            temas TEXT,
            desarrollo TEXT,
            acuerdos TEXT,
            responsables TEXT,
            observaciones TEXT,
            estado TEXT NOT NULL DEFAULT 'borrador',
            created_by TEXT,
            updated_at TEXT
        );

        CREATE TABLE IF NOT EXISTS votaciones (
            id BIGSERIAL PRIMARY KEY,
            acta_id BIGINT NOT NULL,
            titulo TEXT NOT NULL,
            descripcion TEXT,
            estado TEXT NOT NULL DEFAULT 'abierta',
            created_by TEXT,
            created_at TEXT,
            closed_at TEXT,
            condominio_id BIGINT,
            CONSTRAINT fk_votaciones_acta FOREIGN KEY (acta_id) REFERENCES actas(id)
        );

        CREATE TABLE IF NOT EXISTS votacion_opciones (
            id BIGSERIAL PRIMARY KEY,
            votacion_id BIGINT NOT NULL,
            texto TEXT NOT NULL,
            orden INTEGER NOT NULL DEFAULT 1,
            condominio_id BIGINT,
            CONSTRAINT fk_votacion_opciones_votacion FOREIGN KEY (votacion_id) REFERENCES votaciones(id)
        );

        CREATE TABLE IF NOT EXISTS votacion_votos (
            id BIGSERIAL PRIMARY KEY,
            votacion_id BIGINT NOT NULL,
            opcion_id BIGINT NOT NULL,
            user_id BIGINT NOT NULL,
            parcela_id BIGINT,
            created_at TEXT NOT NULL,
            condominio_id BIGINT,
            CONSTRAINT fk_votacion_votos_votacion FOREIGN KEY (votacion_id) REFERENCES votaciones(id),
            CONSTRAINT fk_votacion_votos_opcion FOREIGN KEY (opcion_id) REFERENCES votacion_opciones(id),
            CONSTRAINT fk_votacion_votos_user FOREIGN KEY (user_id) REFERENCES usuarios(id),
            CONSTRAINT fk_votacion_votos_parcela FOREIGN KEY (parcela_id) REFERENCES parcelas(id),
            CONSTRAINT uq_votacion_user UNIQUE (votacion_id, user_id)
        );

        CREATE INDEX IF NOT EXISTS idx_movimientos_fecha ON movimientos(fecha);
        CREATE INDEX IF NOT EXISTS idx_votaciones_acta ON votaciones(acta_id);
        CREATE INDEX IF NOT EXISTS idx_votacion_votos_votacion ON votacion_votos(votacion_id);
        CREATE INDEX IF NOT EXISTS idx_pagos_parcelas_mes ON pagos_parcelas(mes);
        CREATE UNIQUE INDEX IF NOT EXISTS idx_parcelas_nombre_curso_unique ON parcelas ((lower(trim(nombre))), (lower(trim(COALESCE(curso, '')))));
        CREATE UNIQUE INDEX IF NOT EXISTS idx_pagos_parcela_mes_unique ON pagos_parcelas(parcela_id, mes);
        """
    else:
        script = """
        CREATE TABLE IF NOT EXISTS actividades (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            fecha TEXT,
            descripcion TEXT
        );

        CREATE TABLE IF NOT EXISTS movimientos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha TEXT NOT NULL,
            tipo TEXT NOT NULL CHECK(tipo IN ('ingreso', 'gasto')),
            concepto TEXT NOT NULL,
            monto REAL NOT NULL CHECK(monto >= 0),
            actividad_id INTEGER,
            parcela_id INTEGER,
            observacion TEXT,
            origen TEXT NOT NULL DEFAULT 'general',
            FOREIGN KEY (actividad_id) REFERENCES actividades(id)
        );

        CREATE TABLE IF NOT EXISTS parcelas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            curso TEXT,
            cuota_mensual REAL NOT NULL DEFAULT 0,
            apoderado TEXT,
            telefono TEXT,
            direccion TEXT,
            observacion_ficha TEXT,
            activo INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS pagos_parcelas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            parcela_id INTEGER NOT NULL,
            fecha TEXT NOT NULL,
            mes TEXT NOT NULL,
            monto REAL NOT NULL CHECK(monto >= 0),
            observacion TEXT,
            movimiento_id INTEGER,
            FOREIGN KEY (parcela_id) REFERENCES parcelas(id),
            FOREIGN KEY (movimiento_id) REFERENCES movimientos(id)
        );

        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            nombre TEXT NOT NULL,
            activo INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS ciclos_cobranza (
            mes TEXT PRIMARY KEY,
            cuota_referencia REAL NOT NULL DEFAULT 0,
            generado_en TEXT NOT NULL,
            generado_por TEXT NOT NULL,
            observacion TEXT
        );

        CREATE TABLE IF NOT EXISTS actas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo TEXT NOT NULL,
            fecha TEXT NOT NULL,
            lugar TEXT,
            hora_inicio TEXT,
            hora_termino TEXT,
            asistentes TEXT,
            temas TEXT,
            desarrollo TEXT,
            acuerdos TEXT,
            responsables TEXT,
            observaciones TEXT,
            estado TEXT NOT NULL DEFAULT 'borrador',
            created_by TEXT,
            updated_at TEXT
        );

        CREATE TABLE IF NOT EXISTS votaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            acta_id INTEGER NOT NULL,
            titulo TEXT NOT NULL,
            descripcion TEXT,
            estado TEXT NOT NULL DEFAULT 'abierta',
            created_by TEXT,
            created_at TEXT,
            closed_at TEXT,
            condominio_id INTEGER,
            FOREIGN KEY (acta_id) REFERENCES actas(id)
        );

        CREATE TABLE IF NOT EXISTS votacion_opciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            votacion_id INTEGER NOT NULL,
            texto TEXT NOT NULL,
            orden INTEGER NOT NULL DEFAULT 1,
            condominio_id INTEGER,
            FOREIGN KEY (votacion_id) REFERENCES votaciones(id)
        );

        CREATE TABLE IF NOT EXISTS votacion_votos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            votacion_id INTEGER NOT NULL,
            opcion_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            parcela_id INTEGER,
            created_at TEXT NOT NULL,
            condominio_id INTEGER,
            FOREIGN KEY (votacion_id) REFERENCES votaciones(id),
            FOREIGN KEY (opcion_id) REFERENCES votacion_opciones(id),
            FOREIGN KEY (user_id) REFERENCES usuarios(id),
            FOREIGN KEY (parcela_id) REFERENCES parcelas(id),
            UNIQUE (votacion_id, user_id)
        );

        CREATE INDEX IF NOT EXISTS idx_movimientos_fecha ON movimientos(fecha);
        CREATE INDEX IF NOT EXISTS idx_votaciones_acta ON votaciones(acta_id);
        CREATE INDEX IF NOT EXISTS idx_votacion_votos_votacion ON votacion_votos(votacion_id);
        CREATE INDEX IF NOT EXISTS idx_pagos_parcelas_mes ON pagos_parcelas(mes);
        CREATE UNIQUE INDEX IF NOT EXISTS idx_parcelas_nombre_curso_unique ON parcelas(lower(trim(nombre)), lower(trim(COALESCE(curso, ''))));
        CREATE UNIQUE INDEX IF NOT EXISTS idx_pagos_parcela_mes_unique ON pagos_parcelas(parcela_id, mes);
        """
    db.executescript(script)
    migrate_legacy_parcelas_schema(db)
    try:
        db.execute('CREATE TABLE IF NOT EXISTS condominios (id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT NOT NULL UNIQUE, direccion TEXT, activo INTEGER NOT NULL DEFAULT 1)' if db.kind == 'sqlite' else 'CREATE TABLE IF NOT EXISTS condominios (id BIGSERIAL PRIMARY KEY, nombre TEXT NOT NULL UNIQUE, direccion TEXT, activo INTEGER NOT NULL DEFAULT 1)')
        db.commit()
    except Exception:
        db.rollback()

    # Migraciones suaves para bases existentes creadas con versiones anteriores.
    if db.kind == 'postgres':
        migration_statements = [
            'ALTER TABLE actividades ADD COLUMN IF NOT EXISTS descripcion TEXT',
            'ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS actividad_id BIGINT',
            'ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS parcela_id BIGINT',
            "ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS origen TEXT NOT NULL DEFAULT 'general'",
            'ALTER TABLE pagos_parcelas ADD COLUMN IF NOT EXISTS observacion TEXT',
            'ALTER TABLE pagos_parcelas ADD COLUMN IF NOT EXISTS movimiento_id BIGINT',
            'ALTER TABLE parcelas ADD COLUMN IF NOT EXISTS apoderado TEXT',
            'ALTER TABLE parcelas ADD COLUMN IF NOT EXISTS telefono TEXT',
            'ALTER TABLE parcelas ADD COLUMN IF NOT EXISTS direccion TEXT',
            'ALTER TABLE parcelas ADD COLUMN IF NOT EXISTS observacion_ficha TEXT',
            'ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS condominio_id BIGINT',
            'ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS parcela_id BIGINT',
            'ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS parcela_id BIGINT',
            'ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS must_change_password INTEGER NOT NULL DEFAULT 0',
            'ALTER TABLE parcelas ADD COLUMN IF NOT EXISTS condominio_id BIGINT',
            'ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS condominio_id BIGINT',
            'ALTER TABLE pagos_parcelas ADD COLUMN IF NOT EXISTS condominio_id BIGINT',
            'ALTER TABLE actas ADD COLUMN IF NOT EXISTS condominio_id BIGINT',
            'ALTER TABLE actividades ADD COLUMN IF NOT EXISTS condominio_id BIGINT',
            'ALTER TABLE ciclos_cobranza ADD COLUMN IF NOT EXISTS condominio_id BIGINT',
        ]
    else:
        migration_statements = [
            'ALTER TABLE actividades ADD COLUMN descripcion TEXT',
            'ALTER TABLE movimientos ADD COLUMN actividad_id INTEGER',
            'ALTER TABLE movimientos ADD COLUMN parcela_id INTEGER',
            "ALTER TABLE movimientos ADD COLUMN origen TEXT NOT NULL DEFAULT 'general'",
            'ALTER TABLE pagos_parcelas ADD COLUMN observacion TEXT',
            'ALTER TABLE pagos_parcelas ADD COLUMN movimiento_id INTEGER',
            'ALTER TABLE parcelas ADD COLUMN apoderado TEXT',
            'ALTER TABLE parcelas ADD COLUMN telefono TEXT',
            'ALTER TABLE parcelas ADD COLUMN direccion TEXT',
            'ALTER TABLE parcelas ADD COLUMN observacion_ficha TEXT',
            'ALTER TABLE usuarios ADD COLUMN condominio_id INTEGER',
            'ALTER TABLE usuarios ADD COLUMN parcela_id INTEGER',
            'ALTER TABLE usuarios ADD COLUMN must_change_password INTEGER NOT NULL DEFAULT 0',
            'ALTER TABLE parcelas ADD COLUMN condominio_id INTEGER',
            'ALTER TABLE movimientos ADD COLUMN condominio_id INTEGER',
            'ALTER TABLE pagos_parcelas ADD COLUMN condominio_id INTEGER',
            'ALTER TABLE actas ADD COLUMN condominio_id INTEGER',
            'ALTER TABLE actividades ADD COLUMN condominio_id INTEGER',
            'ALTER TABLE ciclos_cobranza ADD COLUMN condominio_id INTEGER',
            'ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS condominio_id BIGINT',
            'ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS parcela_id BIGINT',
            'ALTER TABLE parcelas ADD COLUMN IF NOT EXISTS condominio_id BIGINT',
            'ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS condominio_id BIGINT',
            'ALTER TABLE pagos_parcelas ADD COLUMN IF NOT EXISTS condominio_id BIGINT',
            'ALTER TABLE actas ADD COLUMN IF NOT EXISTS condominio_id BIGINT',
            'ALTER TABLE actividades ADD COLUMN IF NOT EXISTS condominio_id BIGINT',
            'ALTER TABLE ciclos_cobranza ADD COLUMN IF NOT EXISTS condominio_id BIGINT',
        ]
        try:
            db.execute('CREATE TABLE IF NOT EXISTS ciclos_cobranza (mes TEXT PRIMARY KEY, cuota_referencia REAL NOT NULL DEFAULT 0, generado_en TEXT NOT NULL, generado_por TEXT NOT NULL, observacion TEXT)')
            db.commit()
        except Exception:
            db.rollback()
        try:
            db.execute("CREATE TABLE IF NOT EXISTS actas (id INTEGER PRIMARY KEY AUTOINCREMENT, titulo TEXT NOT NULL, fecha TEXT NOT NULL, lugar TEXT, hora_inicio TEXT, hora_termino TEXT, asistentes TEXT, temas TEXT, desarrollo TEXT, acuerdos TEXT, responsables TEXT, observaciones TEXT, estado TEXT NOT NULL DEFAULT 'borrador', created_by TEXT, updated_at TEXT)")
            db.commit()
        except Exception:
            db.rollback()
        try:
            db.execute("CREATE TABLE IF NOT EXISTS votaciones (id INTEGER PRIMARY KEY AUTOINCREMENT, acta_id INTEGER NOT NULL, titulo TEXT NOT NULL, descripcion TEXT, estado TEXT NOT NULL DEFAULT 'abierta', created_by TEXT, created_at TEXT, closed_at TEXT, condominio_id INTEGER)")
            db.commit()
        except Exception:
            db.rollback()
        try:
            db.execute("CREATE TABLE IF NOT EXISTS votacion_opciones (id INTEGER PRIMARY KEY AUTOINCREMENT, votacion_id INTEGER NOT NULL, texto TEXT NOT NULL, orden INTEGER NOT NULL DEFAULT 1, condominio_id INTEGER)")
            db.commit()
        except Exception:
            db.rollback()
        try:
            db.execute("CREATE TABLE IF NOT EXISTS votacion_votos (id INTEGER PRIMARY KEY AUTOINCREMENT, votacion_id INTEGER NOT NULL, opcion_id INTEGER NOT NULL, user_id INTEGER NOT NULL, parcela_id INTEGER, created_at TEXT NOT NULL, condominio_id INTEGER, UNIQUE(votacion_id, user_id))")
            db.commit()
        except Exception:
            db.rollback()
        try:
            create_sql = db.fetchone("SELECT sql FROM sqlite_master WHERE type='table' AND name='usuarios'")
            sql_text = (create_sql['sql'] if create_sql else '') or ''
            if 'solo_lectura' in sql_text:
                db.executescript("""
                ALTER TABLE usuarios RENAME TO usuarios_old;
                CREATE TABLE usuarios (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT NOT NULL UNIQUE,
                    password_hash TEXT NOT NULL,
                    role TEXT NOT NULL,
                    nombre TEXT NOT NULL,
                    activo INTEGER NOT NULL DEFAULT 1,
                    must_change_password INTEGER NOT NULL DEFAULT 0
                );
                INSERT INTO usuarios (id, username, password_hash, role, nombre, activo, must_change_password)
                SELECT id, username, password_hash, CASE WHEN role='solo_lectura' THEN 'comite' ELSE role END, nombre, activo, 0 FROM usuarios_old;
                DROP TABLE usuarios_old;
                """)
                db.commit()
        except Exception:
            db.rollback()

    for statement in migration_statements:
        try:
            db.execute(statement)
            db.commit()
        except Exception:
            db.rollback()
    try:
        default_condominio_id = get_default_condominio_id(db)
        db.execute('UPDATE parcelas SET condominio_id = ? WHERE condominio_id IS NULL', (default_condominio_id,))
        db.execute('UPDATE movimientos SET condominio_id = ? WHERE condominio_id IS NULL', (default_condominio_id,))
        db.execute('UPDATE pagos_parcelas SET condominio_id = ? WHERE condominio_id IS NULL', (default_condominio_id,))
        db.execute('UPDATE actas SET condominio_id = ? WHERE condominio_id IS NULL', (default_condominio_id,))
        db.execute('UPDATE actividades SET condominio_id = ? WHERE condominio_id IS NULL', (default_condominio_id,))
        db.execute('UPDATE ciclos_cobranza SET condominio_id = ? WHERE condominio_id IS NULL', (default_condominio_id,))
        db.commit()
    except Exception:
        db.rollback()
    try:
        db.execute('DROP INDEX IF EXISTS idx_parcelas_nombre_curso_unique')
        db.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_parcelas_nombre_curso_unique ON parcelas(condominio_id, lower(trim(nombre)), lower(trim(COALESCE(curso, ""))))')
        db.commit()
    except Exception:
        db.rollback()
    db.commit()


app = create_app()

if __name__ == '__main__':
    host = os.environ.get('APP_HOST', '0.0.0.0')
    port = int(os.environ.get('PORT', os.environ.get('APP_PORT', '10000')))
    debug = os.environ.get('APP_DEBUG', '0') == '1'
    app.run(host=host, port=port, debug=debug, use_reloader=False)
